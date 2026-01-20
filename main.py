import os
import json
import unicodedata
import re
import traceback
import math
import calendar
from copy import copy
from datetime import datetime, timedelta, date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl.formula.translate import Translator

# =========================
# OPTIONAL: feriados auto
# =========================
try:
    import holidays as pyholidays
except Exception:
    pyholidays = None


# ============================================================
# CONFIG
# ============================================================
DEFAULT_CONFIG = {
    "company_name": "TALCA",
    "default_jornada_weekday": 8,
    "employees_excel_path": "",
    "template_excel_path": "",
    # feriados (se guardan como YYYY-MM-DD)
    "holidays": [],
    # auto feriados
    "auto_holidays_enabled": True,
    "auto_holidays_country": "AR",
    "auto_holidays_subdiv": "M",  # Mendoza
    "auto_holidays_observed": True,
    # debug opcional
    "debug": False,
    # modo merge si un empleado ya estaba:
    # "sum" suma horas en la celda
    # "replace" pisa horas en la celda
    "merge_mode": "sum",

    # --- IMPORTANTE ---
    # Dejo estos por compatibilidad (ya no los usamos para calcular extra, porque ahora:
    # - Lun-Vie: extra = horas_trab - jornada
    # - Sab/Dom/Feriado: todo lo trabajado cuenta como extra
    "saturday_standard_hours": 0,
    "sunday_standard_hours": 0,
    "holiday_standard_hours": 0,
}

CONFIG_FILE = "config_horas_extra.json"
EMP_MASTER_DEFAULT_NAME = "datos empleados.xlsx"
TEMPLATE_DEFAULT_NAME = "formatosugerido.xlsx"

TEMPLATE_SHEET_NAME = "_TEMPLATE"  # quedará OCULTA en el Excel final

DOW_MAP = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"}


# ============================================================
# HELPERS
# ============================================================
def normalize_text(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFKD", s)
        if not unicodedata.combining(c)
    )
    # saca comas, puntos, guiones, etc. (deja letras/numeros/espacios)
    s = re.sub(r"[^a-z0-9\s]+", " ", s)
    s = " ".join(s.split())
    return s


def name_keys(raw: str):
    """
    Devuelve 3 claves para matchear nombres aunque vengan con extras:
    - full: nombre normalizado completo
    - first2: primeras 2 palabras
    - last2: últimas 2 palabras
    """
    full = normalize_text(raw)
    toks = full.split()
    first2 = " ".join(toks[:2]) if len(toks) >= 2 else full
    last2 = " ".join(toks[-2:]) if len(toks) >= 2 else full
    return full, first2, last2


def load_config() -> dict:
    if not os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_CONFIG, f, ensure_ascii=False, indent=2)

    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    for k, v in DEFAULT_CONFIG.items():
        if k not in cfg:
            cfg[k] = v
    return cfg


def save_config(cfg: dict) -> None:
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def debug_enabled(cfg: dict) -> bool:
    return bool(cfg.get("debug", False))


def debug_write_text(cfg: dict, filename: str, content: str) -> None:
    if not debug_enabled(cfg):
        return
    with open(filename, "w", encoding="utf-8") as f:
        f.write(content)


def debug_write_df(cfg: dict, filename: str, df: pd.DataFrame) -> None:
    if not debug_enabled(cfg):
        return
    df.to_csv(filename, index=False, encoding="utf-8-sig")


def clean_id(val) -> str:
    if pd.isna(val):
        return ""
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float) and float(val).is_integer():
        return str(int(val))
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def only_digits(s: str) -> str:
    return re.sub(r"\D", "", str(s or ""))


def id_key_from_any(val) -> str:
    """
    Clave interna estable para matchear:
    - deja solo dígitos
    - quita ceros a la izquierda (00123 == 123)
    """
    digits = only_digits(clean_id(val))
    if not digits:
        return ""
    k = digits.lstrip("0")
    return k if k else "0"


def extract_id_parts(val):
    """
    Devuelve:
      - key: clave estable (sin ceros a la izquierda)
      - digits: dígitos completos
      - cuil11: si parece CUIL (11)
      - dni8: si se puede inferir DNI (8) (desde DNI o desde CUIL)
    """
    digits = only_digits(clean_id(val))
    key = id_key_from_any(val)

    cuil11 = digits if len(digits) == 11 else ""

    dni8 = ""
    if len(digits) == 11:
        dni8 = digits[2:10].zfill(8)
    elif len(digits) == 8:
        dni8 = digits
    elif len(digits) == 7:
        dni8 = digits.zfill(8)
    elif len(digits) > 8:
        dni8 = digits[-8:]

    return key, digits, cuil11, dni8


def parse_date(val):
    if pd.isna(val):
        return None

    if isinstance(val, (int, float)) and val > 30000:
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(val))).date()

    if isinstance(val, (pd.Timestamp, datetime)):
        return val.date()

    s = str(val).strip()
    m = re.search(r'(\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4})', s)
    if m:
        s = m.group(1)

    s = s.replace(".", "/").replace("-", "/")
    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return None
    return dt.date()


def parse_time_only(val):
    if pd.isna(val):
        return None

    if isinstance(val, (pd.Timestamp, datetime)):
        return val.time()

    if isinstance(val, (int, float)) and 0 <= float(val) < 1:
        total_seconds = int(round(float(val) * 24 * 3600))
        hh = total_seconds // 3600
        mm = (total_seconds % 3600) // 60
        ss = total_seconds % 60
        return datetime(2000, 1, 1, hh, mm, ss).time()

    s = str(val).strip()
    if not s:
        return None

    dt = pd.to_datetime(s, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.time()

def round_dt_to_nearest_hour(dt: datetime) -> datetime:
    """
    Redondea al entero de hora más cercano:
    - 20:58 -> 21:00
    - 06:01 -> 06:00
    - 06:05 -> 06:00
    - 18:10 -> 18:00
    Regla: suma 30 min y luego trunca a la hora.
    """
    return (dt + timedelta(minutes=30)).replace(minute=0, second=0, microsecond=0)



def floor_time_to_hour(t):
    """Trunca minutos/segundos: 06:05 -> 06:00"""
    if t is None:
        return None
    return t.replace(minute=0, second=0, microsecond=0)

def split_interval_by_date(start_dt: datetime, end_dt: datetime):
    """
    Parte un intervalo en horas por cada fecha calendario.
    Devuelve dict: {date: horas_int}
    """
    out = {}
    cur = start_dt
    while cur.date() < end_dt.date():
        midnight = datetime.combine(cur.date() + timedelta(days=1), datetime.min.time())
        hs = int((midnight - cur).total_seconds() // 3600)
        out[cur.date()] = out.get(cur.date(), 0) + hs
        cur = midnight

    hs_last = int((end_dt - cur).total_seconds() // 3600)
    out[end_dt.date()] = out.get(end_dt.date(), 0) + hs_last
    return out


def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())


def week_key_sun_to_sat(d: date) -> date:
    # domingo pertenece a la semana que arranca el lunes siguiente
    if d.weekday() == 6:
        d = d + timedelta(days=1)
    return monday_of_week(d)


def hours_between(start_dt: datetime, end_dt: datetime) -> float:
    if end_dt < start_dt:
        end_dt = end_dt + timedelta(days=1)
    return (end_dt - start_dt).total_seconds() / 3600.0


def is_holiday(d: date, cfg: dict) -> bool:
    hol = set(cfg.get("holidays", []))
    return d.strftime("%Y-%m-%d") in hol


def guess_col(columns_norm, keywords):
    for i, c in enumerate(columns_norm):
        for k in keywords:
            if k in c:
                return i
    return None


def excel_cell_to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        dt = pd.to_datetime(str(v), errors="coerce", dayfirst=True)
        if pd.isna(dt):
            return None
        return dt.date()
    except Exception:
        return None


def round_hours(x: float, step: float = 0.25) -> float:
    """
    Redondeo de horas a "pasos" (por defecto 15 min = 0.25h).
    Esto evita cosas tipo 0.49999997 por floats.
    """
    try:
        x = float(x)
    except Exception:
        return 0.0
    if x <= 0:
        return 0.0
    # redondeo al step más cercano (sin banker's rounding)
    return round(math.floor(x / step + 0.5) * step, 2)



def num_or_zero(v):
    try:
        if v is None or v == "":
            return 0.0
        return float(v)
    except Exception:
        return 0.0

def force_integer_hours_format(ws, row, col_start=5, col_end=12):
    """
    Fuerza que las celdas de horas (Dom..Feriado = columnas 5..12)
    se vean como enteros (sin ,00).
    """
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row, c)

        # convierte el valor a int si hay algo
        try:
            if cell.value not in (None, ""):
                cell.value = int(float(cell.value))
        except Exception:
            pass

        # formato de número sin decimales
        cell.number_format = "0"
        cell.alignment = Alignment(horizontal="center", vertical="center")


def compute_amounts_for_row(ws, row, rate_lav, rate_sab, rate_domfer):
    # Columnas horas (según tu plantilla):
    # Dom=5, Lun=6, Mar=7, Mier=8, Jue=9, Vie=10, Sab=11, Feriado=12
    hours_dom = num_or_zero(ws.cell(row, 5).value)
    hours_lun = num_or_zero(ws.cell(row, 6).value)
    hours_mar = num_or_zero(ws.cell(row, 7).value)
    hours_mie = num_or_zero(ws.cell(row, 8).value)
    hours_jue = num_or_zero(ws.cell(row, 9).value)
    hours_vie = num_or_zero(ws.cell(row, 10).value)
    hours_sab = num_or_zero(ws.cell(row, 11).value)
    hours_fer = num_or_zero(ws.cell(row, 12).value)

    hours_weekday = hours_lun + hours_mar + hours_mie + hours_jue + hours_vie
    hours_domfer = hours_dom + hours_fer

    amt_weekday = int(round(hours_weekday * float(rate_lav), 0))
    amt_sat = int(round(hours_sab * float(rate_sab), 0))
    amt_domfer = int(round(hours_domfer * float(rate_domfer), 0))

    return amt_weekday, amt_sat, amt_domfer


# ============================================================
# AUTO HOLIDAYS (Argentina - Mendoza)
# ============================================================
def compute_auto_holidays_for_range(start: date, end: date, cfg: dict):
    if pyholidays is None:
        return []
    if start is None or end is None:
        return []

    years = list(range(start.year, end.year + 1))
    country = cfg.get("auto_holidays_country", "AR")
    subdiv = cfg.get("auto_holidays_subdiv", "M")
    observed = bool(cfg.get("auto_holidays_observed", True))

    try:
        hcal = pyholidays.country_holidays(country, subdiv=subdiv, years=years, observed=observed)
    except Exception:
        return []

    out = []
    cur = start
    while cur <= end:
        if cur in hcal:
            out.append(cur)
        cur += timedelta(days=1)
    return out


# ============================================================
# MASTER EMPLEADOS
# ============================================================
def load_employee_master(emp_path: str) -> pd.DataFrame:
    df = pd.read_excel(emp_path)
    df.columns = [str(c).strip() for c in df.columns]
    colmap = {normalize_text(c): c for c in df.columns}

    def col_like(keys):
        for k in keys:
            for norm, orig in colmap.items():
                if k in norm:
                    return orig
        return None

    c_empresa = col_like(["empresa"])
    c_sector = col_like(["sector"])
    c_id = col_like(["id"])  # en tu master se llama ID
    c_name = col_like(["apellido", "nombre"])
    c_jornada = col_like(["jornada"])
    c_lav = col_like(["lav"])
    c_sab = col_like(["sab"])
    c_domfer = col_like(["domyfer", "dom", "fer"])

    need = [c_empresa, c_sector, c_id, c_name, c_jornada, c_lav, c_sab, c_domfer]
    if any(x is None for x in need):
        raise RuntimeError(
            "En datos empleados.xlsx faltan columnas esperadas.\n"
            "Necesito: EMPRESA, SECTOR, ID, APELLIDO Y NOMBRE, JORNADA, $/hs LaV, $/hs Sab, $/hs DomYFer."
        )

    out = pd.DataFrame()
    out["id_master_display"] = df[c_id].apply(clean_id).astype(str).str.strip()

    parts = out["id_master_display"].apply(lambda x: pd.Series(
        extract_id_parts(x),
        index=["id_key", "id_digits", "id_cuil11", "id_dni8"]
    ))
    out = pd.concat([out, parts], axis=1)

    out["empresa_master"] = df[c_empresa].fillna("").astype(str).str.strip()
    out["sector_master"] = df[c_sector].fillna("").astype(str).str.strip().apply(normalize_text)

    out["nombre_master"] = df[c_name].fillna("").astype(str).str.strip()
    nk = out["nombre_master"].apply(lambda s: pd.Series(name_keys(s), index=["nombre_norm", "nombre_first2", "nombre_last2"]))
    out = pd.concat([out, nk], axis=1)

    out["jornada_weekday"] = pd.to_numeric(df[c_jornada], errors="coerce").fillna(0).astype(float)
    out["rate_lav"] = pd.to_numeric(df[c_lav], errors="coerce").fillna(0).astype(float)
    out["rate_sab"] = pd.to_numeric(df[c_sab], errors="coerce").fillna(0).astype(float)
    out["rate_domfer"] = pd.to_numeric(df[c_domfer], errors="coerce").fillna(0).astype(float)

    out = out[out["id_key"] != ""].copy()
    out = out.reset_index(drop=True)
    out["midx"] = out.index  # id interno estable

    return out


# ============================================================
# LECTURA REPORTE VEOTIME
# ============================================================
def read_report_raw(path: str) -> pd.DataFrame:
    with open(path, "rb") as f:
        head = f.read(8192).lstrip()
    head_low = head.lower()

    # xlsx
    if head.startswith(b"PK"):
        return pd.read_excel(path, engine="openpyxl", header=None)

    # xls real (OLE)
    if head.startswith(b"\xD0\xCF\x11\xE0"):
        return pd.read_excel(path, engine="xlrd", header=None)

    # html disfrazado
    if b"<html" in head_low or b"<!doctype" in head_low or b"<table" in head_low:
        tables = pd.read_html(path)

        def score_table(t: pd.DataFrame) -> int:
            txt = " ".join([str(x) for x in list(t.columns)] + t.astype(str).values.ravel().tolist())
            txt = normalize_text(txt)
            score = 0
            for kw in ["fecha", "hora", "dni", "marc", "entrada", "salida"]:
                if kw in txt:
                    score += 1
            return score

        scored = [(score_table(t), t) for t in tables]
        scored.sort(key=lambda x: x[0], reverse=True)
        best = scored[0][1].copy()
        best.columns = list(range(best.shape[1]))
        return best.reset_index(drop=True)

    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return pd.read_excel(path, engine="openpyxl", header=None)
    if ext == ".xls":
        return pd.read_excel(path, engine="xlrd", header=None)

    raise RuntimeError("Formato no soportado. Usá .xls o .xlsx.")


def find_header_row(df_raw, max_rows=80):
    for r in range(min(max_rows, len(df_raw))):
        row = df_raw.iloc[r].astype(str).tolist()
        text = " ".join(normalize_text(x) for x in row)

        has_fecha = "fecha" in text
        has_hora = "hora" in text or "time" in text
        has_id = ("dni" in text) or ("documento" in text) or ("legajo" in text) or ("id" in text)
        has_marc = "marc" in text or "tipo" in text

        if has_fecha and has_hora and has_id and has_marc:
            return r
    return None


def apply_header_row(df_raw, header_row):
    headers = df_raw.iloc[header_row].tolist()
    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = [str(h).strip() for h in headers]
    df = df.dropna(axis=1, how="all")
    return df


def read_veotime_to_daily(path: str, emp_master: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    df_raw = read_report_raw(path)

    header_row = find_header_row(df_raw)
    if header_row is not None:
        df = apply_header_row(df_raw, header_row)
    else:
        df = df_raw.copy()
        df.columns = [str(c).strip() for c in df.iloc[0].tolist()]
        df = df.iloc[1:].copy()

    cols_norm = [normalize_text(c) for c in df.columns]

    i_fecha = guess_col(cols_norm, ["fecha"])
    i_hora = guess_col(cols_norm, ["hora", "time"])
    i_marc = guess_col(cols_norm, ["marcaci", "marc", "tipo"])
    i_id = guess_col(cols_norm, ["dni", "documento", "legajo", "id"])
    i_nombre = guess_col(cols_norm, ["nombre", "empleado", "apellido", "colaborador"])

    missing = []
    if i_fecha is None: missing.append("Fecha")
    if i_hora is None: missing.append("Hora")
    if i_marc is None: missing.append("Marcación")
    if i_id is None: missing.append("DNI/ID/Legajo")
    if i_nombre is None: missing.append("Nombre")

    if missing:
        debug_write_df(cfg, "debug_veotime_head.csv", df.head(200))
        debug_write_text(cfg, "debug_columnas.txt", "Columnas detectadas:\n" + "\n".join(f"- {c}" for c in df.columns))
        raise RuntimeError(
            "No pude detectar columnas clave en el reporte VeoTime.\n"
            f"Faltan: {', '.join(missing)}"
        )

    def norm_tipo(x):
        t = normalize_text(x)
        if "entrada" in t or "ingreso" in t or t.startswith("ent"):
            return "entrada"
        if "salida" in t or "egreso" in t or t.startswith("sal"):
            return "salida"
        return ""

    events = pd.DataFrame()
    events["fecha"] = df.iloc[:, i_fecha].apply(parse_date)
    events["hora"] = df.iloc[:, i_hora].apply(parse_time_only)
    events["tipo"] = df.iloc[:, i_marc].apply(norm_tipo)

    events["raw_id_rep"] = df.iloc[:, i_id].apply(clean_id).astype(str).str.strip()
    parts = events["raw_id_rep"].apply(lambda x: pd.Series(
        extract_id_parts(x),
        index=["id_key_rep", "id_digits_rep", "id_cuil11_rep", "id_dni8_rep"]
    ))
    events = pd.concat([events, parts], axis=1)

    events["nombre_rep"] = df.iloc[:, i_nombre].astype(str).str.strip()
    nk = events["nombre_rep"].apply(lambda s: pd.Series(name_keys(s), index=["nombre_norm_rep", "nombre_first2_rep", "nombre_last2_rep"]))
    events = pd.concat([events, nk], axis=1)

    # filtros
    events = events.dropna(subset=["fecha", "hora"])
    events = events[events["tipo"].isin(["entrada", "salida"])]
    events = events[events["id_key_rep"] != ""]

    if events.empty:
        debug_write_df(cfg, "debug_veotime_head.csv", df.head(200))
        raise RuntimeError("No quedaron eventos válidos (Entrada/Salida).")

    emp = emp_master.copy()

    # ---- mapeos por CLAVES ÚNICAS ----
    def unique_map(key_col):
        s = emp[[key_col, "midx"]].copy()
        s[key_col] = s[key_col].fillna("").astype(str).str.strip()
        s = s[s[key_col] != ""]
        if s.empty:
            return {}
        vc = s[key_col].value_counts()
        uniques = set(vc[vc == 1].index.astype(str))
        s = s[s[key_col].isin(uniques)]
        return dict(zip(s[key_col].astype(str), s["midx"]))

    map_digits = unique_map("id_digits")
    map_cuil = unique_map("id_cuil11")
    map_dni8  = unique_map("id_dni8")
    map_key   = unique_map("id_key")

    map_name_full  = unique_map("nombre_norm")
    map_name_f2    = unique_map("nombre_first2")
    map_name_l2    = unique_map("nombre_last2")

    # ---- matching por ID primero ----
    events["midx"] = events["id_digits_rep"].astype(str).map(map_digits)

    mask = events["midx"].isna()
    events.loc[mask, "midx"] = events.loc[mask, "id_cuil11_rep"].astype(str).map(map_cuil)

    mask = events["midx"].isna()
    events.loc[mask, "midx"] = events.loc[mask, "id_dni8_rep"].astype(str).map(map_dni8)

    mask = events["midx"].isna()
    events.loc[mask, "midx"] = events.loc[mask, "id_key_rep"].astype(str).map(map_key)

    # ---- si no matcheó por ID, matcheo por nombre (robusto) ----
    mask = events["midx"].isna()
    events.loc[mask, "midx"] = events.loc[mask, "nombre_norm_rep"].astype(str).map(map_name_full)

    mask = events["midx"].isna()
    events.loc[mask, "midx"] = events.loc[mask, "nombre_first2_rep"].astype(str).map(map_name_f2)

    mask = events["midx"].isna()
    events.loc[mask, "midx"] = events.loc[mask, "nombre_last2_rep"].astype(str).map(map_name_l2)

    merged = events.merge(emp, how="left", on="midx")

    still_no = merged["nombre_master"].isna()
    if still_no.any():
        debug_write_df(cfg, "debug_no_matcheados.csv",
                       merged.loc[still_no, ["raw_id_rep", "id_key_rep", "nombre_rep"]].drop_duplicates())

    # finales
    merged["empresa_final"] = merged["empresa_master"].fillna("").astype(str).str.strip()
    merged["empresa_final"] = merged["empresa_final"].where(merged["empresa_final"] != "", cfg.get("company_name", "TALCA"))

    merged["sector_final"] = merged["sector_master"].fillna("").astype(str).str.strip()
    merged["sector_final"] = merged["sector_final"].where(merged["sector_final"] != "", "desconocido")
    merged["sector_final"] = merged["sector_final"].apply(normalize_text)

    merged["nombre_final"] = merged["nombre_master"].fillna("").astype(str).str.strip()
    merged["nombre_final"] = merged["nombre_final"].where(merged["nombre_final"] != "", merged["nombre_rep"].fillna("").astype(str))

    merged["jornada_weekday"] = merged["jornada_weekday"].fillna(float(cfg.get("default_jornada_weekday", 8)))
    merged.loc[merged["jornada_weekday"] <= 0, "jornada_weekday"] = float(cfg.get("default_jornada_weekday", 8))

    merged["rate_lav"] = merged["rate_lav"].fillna(0.0)
    merged["rate_sab"] = merged["rate_sab"].fillna(0.0)
    merged["rate_domfer"] = merged["rate_domfer"].fillna(0.0)

    # ID final: escribe SIEMPRE el ID del master si existe
    merged["id_display_final"] = merged["id_master_display"].fillna("").astype(str).str.strip()
    merged["id_display_final"] = merged["id_display_final"].where(
        merged["id_display_final"] != "", merged["raw_id_rep"].fillna("").astype(str).str.strip()
    )

    # clave interna estable para agrupar
    merged["id_key_final"] = merged["id_key"].fillna("").astype(str).str.strip()
    merged["id_key_final"] = merged["id_key_final"].where(
        merged["id_key_final"] != "", merged["id_key_rep"].fillna("").astype(str).str.strip()
    )

    # =========================================================
    # CONSTRUIR HORAS DIARIAS (TURNO NOCTURNO)
    # - Redondeo a hora más cercana (no minutos)
    # - Empareja Entrada -> Salida aunque sea al día siguiente
    # - Asigna TODAS las horas al día de la ENTRADA (como hace RRHH)
    # - Si cruza medianoche: suma +1 hora extra (bonus nocturno) al día de entrada
    # =========================================================

    # datetime crudo (fecha + hora)
    merged["dt_raw"] = merged.apply(
        lambda r: datetime.combine(r["fecha"], r["hora"]) if (
                    r["fecha"] is not None and r["hora"] is not None) else None,
        axis=1
    )
    merged = merged.dropna(subset=["dt_raw"]).copy()

    # redondeo a la hora mas cercana
    merged["dt"] = merged["dt_raw"].apply(round_dt_to_nearest_hour)

    # acumulador por (dni, fecha_asignada)
    acc = {}

    def get_rec(static, dte: date):
        k = (static["dni"], dte)
        if k not in acc:
            acc[k] = {
                "dni": static["dni"],
                "dni_display": static["dni_display"],
                "empresa": static["empresa"],
                "nombre": static["nombre"],
                "sector": static["sector"],
                "fecha": dte,
                "horas_trab": 0,  # horas enteras ya redondeadas
                "night_bonus": 0,  # +1 si cruzo medianoche
                "jornada_weekday": static["jornada_weekday"],
                "rate_lav": static["rate_lav"],
                "rate_sab": static["rate_sab"],
                "rate_domfer": static["rate_domfer"],
            }
        return acc[k]

    # agrupamos por empleado (sin fecha), así emparejamos turnos cruzando días
    emp_group_cols = [
        "id_key_final", "id_display_final",
        "empresa_final", "nombre_final", "sector_final",
        "jornada_weekday", "rate_lav", "rate_sab", "rate_domfer"
    ]

    for key, sub in merged.groupby(emp_group_cols, dropna=False):
        (idkey, iddisp, empresa, nombre, sector, jornada, rlav, rsab, rdom) = key

        static = {
            "dni": str(idkey).strip(),
            "dni_display": str(iddisp).strip() if str(iddisp).strip() else str(idkey).strip(),
            "empresa": str(empresa),
            "nombre": str(nombre),
            "sector": str(sector),
            "jornada_weekday": float(jornada),
            "rate_lav": float(rlav),
            "rate_sab": float(rsab),
            "rate_domfer": float(rdom),
        }

        if not static["dni"]:
            continue

        sub = sub.sort_values("dt")

        open_entry = None

        for _, r in sub.iterrows():
            if r["tipo"] == "entrada":
                # si había otra entrada abierta, nos quedamos con la más reciente
                open_entry = r["dt"]
                continue

            if r["tipo"] == "salida" and open_entry is not None:
                end_dt = r["dt"]

                # si por cualquier motivo quedó mal ordenado, lo reacomodamos
                while end_dt <= open_entry:
                    end_dt = end_dt + timedelta(days=1)

                # horas enteras
                hs = int((end_dt - open_entry).total_seconds() // 3600)
                if hs < 0:
                    hs = 0

                day_assigned = open_entry.date()  # TODO va al día de la entrada
                rec = get_rec(static, day_assigned)
                rec["horas_trab"] += hs

                # bonus nocturno si cruza medianoche (entrada un día, salida otro)
                if open_entry.date() != end_dt.date():
                    rec["night_bonus"] += 1

                open_entry = None

        # si queda una entrada sin salida -> no cuenta (queda 0), como pediste

    daily = pd.DataFrame(list(acc.values()))
    if daily.empty:
        raise RuntimeError("No pude construir horas diarias (no se formaron pares Entrada/Salida).")
    return daily


# ============================================================
# OVERTIME (CORREGIDO: SIN FLOOR + REGLA TALCA)
# ============================================================
def compute_overtime_from_daily(daily: pd.DataFrame, cfg: dict) -> dict:
    d = daily.copy()
    d["semana_lunes"] = d["fecha"].apply(week_key_sun_to_sat)
    d["dow"] = d["fecha"].apply(lambda x: x.weekday())
    d["dow_key"] = d["dow"].map(DOW_MAP)

    # todo entero
    d["horas_trab"] = pd.to_numeric(d["horas_trab"], errors="coerce").fillna(0).astype(int)
    d["jornada_weekday"] = pd.to_numeric(d["jornada_weekday"], errors="coerce").fillna(0).astype(int)
    d["night_bonus"] = pd.to_numeric(d.get("night_bonus", 0), errors="coerce").fillna(0).astype(int)

    base_extra = (d["horas_trab"] - d["jornada_weekday"]).clip(lower=0).astype(int)
    d["horas_extra"] = (base_extra + d["night_bonus"]).astype(int)

    # tarifa según el día donde se escribe (día asignado = día de entrada)
    def rate_for(row):
        if is_holiday(row["fecha"], cfg) or row["dow_key"] == "Sun":
            return float(row["rate_domfer"])
        if row["dow_key"] == "Sat":
            return float(row["rate_sab"])
        return float(row["rate_lav"])

    d["tarifa_he"] = d.apply(rate_for, axis=1)
    d["costo"] = d["horas_extra"] * d["tarifa_he"]

    weeks = {}
    for week_start, sub in d.groupby("semana_lunes"):
        weeks[week_start] = sub.copy()
    return weeks

# ============================================================
# TEMPLATE HELPERS
# ============================================================
def _safe_sheet_title(s: str) -> str:
    s = re.sub(r'[\\/*?:\[\]]', '-', str(s))
    return s[:31].strip() or "Semana"


def _col_for_date(d: date, cfg: dict) -> int:
    # En tu plantilla: Dom=5, Lun=6, ... Sab=11, Feriado=12
    if is_holiday(d, cfg):
        return 12
    if d.weekday() == 6:
        return 5
    return 6 + d.weekday()


def _ensure_rows(ws, start_row, n_rows_needed, base_style_row=7):
    max_col = max(ws.max_column, 30)
    last_needed = start_row + n_rows_needed - 1
    if last_needed <= ws.max_row:
        return

    for r in range(ws.max_row + 1, last_needed + 1):
        ws.insert_rows(r)
        for c in range(1, max_col + 1):
            src = ws.cell(base_style_row, c)
            dst = ws.cell(r, c)

            dst._style = copy(src._style)
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)

            v = src.value
            if isinstance(v, str) and v.startswith("="):
                try:
                    dst.value = Translator(v, origin=src.coordinate).translate_formula(dst.coordinate)
                except Exception:
                    dst.value = v
            else:
                dst.value = v


def _set_week_header(ws, week_start: date, cfg: dict):
    sun = week_start - timedelta(days=1)
    dates = [
        (5, sun),
        (6, week_start),
        (7, week_start + timedelta(days=1)),
        (8, week_start + timedelta(days=2)),
        (9, week_start + timedelta(days=3)),
        (10, week_start + timedelta(days=4)),
        (11, week_start + timedelta(days=5)),
    ]
    for col, d in dates:
        cell = ws.cell(4, col)
        cell.value = d
        cell.number_format = "dd/mm/yy"

    hol = None
    for i in range(-1, 6):
        d = week_start + timedelta(days=i)
        if is_holiday(d, cfg):
            hol = d
            break
    ws.cell(4, 12).value = hol if hol else None
    ws.cell(4, 12).number_format = "dd/mm/yy"


def _clear_data_area(ws, start_row=7):
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, 5):
            ws.cell(r, c).value = None

        # Horas (E..L) con decimales
        for c in range(5, 13):
            ws.cell(r, c).value = 0
            ws.cell(r, c).number_format = "0.00"
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")

        for c in range(13, 16):
            ws.cell(r, c).value = 0
            ws.cell(r, c).number_format = "#,##0"
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")


# ============================================================
# TEMPLATE: asegurar _TEMPLATE oculta dentro del workbook de salida
# ============================================================
def clone_template_sheet_into_workbook(wb_dest, template_path: str) -> None:
    src_wb = load_workbook(template_path)
    src = src_wb.active

    if TEMPLATE_SHEET_NAME in wb_dest.sheetnames:
        src_wb.close()
        return

    dst = wb_dest.create_sheet(TEMPLATE_SHEET_NAME)

    for col_letter, dim in src.column_dimensions.items():
        dst.column_dimensions[col_letter].width = dim.width

    for row_idx, dim in src.row_dimensions.items():
        dst.row_dimensions[row_idx].height = dim.height

    for row in src.iter_rows():
        for cell in row:
            new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    for m in src.merged_cells.ranges:
        dst.merge_cells(str(m))

    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    dst.page_setup = copy(src.page_setup)
    dst.page_margins = copy(src.page_margins)

    src_wb.close()


def ensure_hidden_template(wb, template_path: str):
    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        if "TEMPLATE" in wb.sheetnames:
            ws = wb["TEMPLATE"]
            ws.title = TEMPLATE_SHEET_NAME

    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        clone_template_sheet_into_workbook(wb, template_path)

    tpl = wb[TEMPLATE_SHEET_NAME]
    tpl.sheet_state = "veryHidden"
    return tpl


def find_week_sheet(wb, week_start: date):
    target_title = _safe_sheet_title(f"Semana {week_start.strftime('%d-%m')}")
    if target_title in wb.sheetnames:
        return wb[target_title]

    for name in wb.sheetnames:
        if name == TEMPLATE_SHEET_NAME:
            continue
        ws = wb[name]
        v = ws.cell(4, 6).value  # F4 = lunes
        d = excel_cell_to_date(v)
        if d == week_start:
            return ws
    return None


def get_or_create_week_sheet(wb, tpl, week_start: date, cfg: dict):
    ws = find_week_sheet(wb, week_start)
    created_new = False

    if ws is None:
        ws = wb.copy_worksheet(tpl)
        ws.sheet_state = "visible"
        ws.title = _safe_sheet_title(f"Semana {week_start.strftime('%d-%m')}")
        _set_week_header(ws, week_start, cfg)
        _clear_data_area(ws, start_row=7)
        created_new = True
    else:
        _set_week_header(ws, week_start, cfg)

    return ws, created_new


def read_existing_employee_rows(ws, start_row=7):
    row_by_key = {}
    last_filled = start_row - 1
    empty_run = 0

    for r in range(start_row, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value

        has_data = any(v not in (None, "") for v in [a, b, c, d])

        if has_data:
            empty_run = 0
            last_filled = r
            k = id_key_from_any(c)  # <-- clave estable aunque tenga guiones/ceros
            if k:
                row_by_key[k] = r
        else:
            empty_run += 1
            if empty_run >= 30 and r > start_row + 30:
                break

    return row_by_key, (last_filled + 1)


def _ensure_row_formulas_from_base(ws, target_row, base_row=7, max_col=30, skip_cols=None):
    skip_cols = set(skip_cols or [])
    for c in range(1, max_col + 1):
        if c in skip_cols:
            continue
        src = ws.cell(base_row, c)
        v = src.value
        if isinstance(v, str) and v.startswith("="):
            dst = ws.cell(target_row, c)
            # si ya hay una formula, no tocar
            if isinstance(dst.value, str) and dst.value.startswith("="):
                continue
            try:
                dst.value = Translator(v, origin=src.coordinate).translate_formula(dst.coordinate)
            except Exception:
                dst.value = v



def upsert_week_employees(ws, df_week: pd.DataFrame, cfg: dict):
    if df_week is None or df_week.empty:
        return

    sub = df_week.copy()

    # agrupar horas extra por empleado y día
    day_he = (
        sub.groupby(
            ["empresa", "dni", "dni_display", "nombre", "sector", "fecha", "rate_lav", "rate_sab", "rate_domfer"],
            dropna=False
        )["horas_extra"]
        .sum()
        .reset_index()
    )

    employees = (
        day_he[["empresa", "dni", "dni_display", "nombre", "sector", "rate_lav", "rate_sab", "rate_domfer"]]
        .drop_duplicates()
        .sort_values(["empresa", "sector", "nombre"])
    )

    # mapa horas extra por key interna y fecha (AHORA float, no int)
    he_map = {}
    for _, r in day_he.iterrows():
        k = str(r["dni"]).strip()
        he_map.setdefault(k, {})[r["fecha"]] = float(r["horas_extra"])

    row_by_key, next_row = read_existing_employee_rows(ws, start_row=7)

    _ensure_rows(ws, start_row=7, n_rows_needed=max(10, next_row - 6 + len(employees)), base_style_row=7)

    merge_mode = str(cfg.get("merge_mode", "sum")).strip().lower()

    for _, emp in employees.iterrows():
        empresa = str(emp["empresa"])
        key_internal = str(emp["dni"]).strip()
        dni_display = str(emp.get("dni_display", "")).strip() or key_internal
        nombre = str(emp["nombre"])
        sector = str(emp["sector"])

        rate_lav = float(emp["rate_lav"])
        rate_sab = float(emp["rate_sab"])
        rate_domfer = float(emp["rate_domfer"])

        if not key_internal:
            continue

        if key_internal in row_by_key:
            row = row_by_key[key_internal]
            is_new = False
        else:
            row = next_row
            next_row += 1
            row_by_key[key_internal] = row
            is_new = True

        # A..D
        ws.cell(row, 1).value = empresa
        ws.cell(row, 2).value = sector
        ws.cell(row, 3).value = dni_display   # <-- SIEMPRE el ID del master si existe
        ws.cell(row, 4).value = nombre

        # Si es fila nueva, inicializamos horas E..L en 0 (con decimales)
        if is_new:
            for c in range(5, 13):
                ws.cell(row, c).value = 0
                ws.cell(row, c).number_format = "0.00"
                ws.cell(row, c).alignment = Alignment(horizontal="center", vertical="center")

        # Horas extra por fecha (E..L)

        emp_he_by_date = he_map.get(key_internal, {})
        for dte, he in emp_he_by_date.items():
            he = float(he)
            if he <= 0:
                continue
            col = _col_for_date(dte, cfg)
            cur = num_or_zero(ws.cell(row, col).value)

            if merge_mode == "replace":
                new_val = he
            else:
                new_val = cur + he

            new_val = round(new_val, 2)
            ws.cell(row, col).value = new_val
            ws.cell(row, col).number_format = "0.00"
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")

        force_integer_hours_format(ws, row, 5, 12)

        # Asegurar fórmulas en la fila
        _ensure_row_formulas_from_base(
            ws,
            target_row=row,
            base_row=7,
            max_col=max(ws.max_column, 30),
            skip_cols={13, 14, 15}  # NO tocar M,N,O porque ahora son montos calculados
        )

        # --- Luego de cargar horas E..L, calculamos montos M..O ---
        amt_weekday, amt_sat, amt_domfer = compute_amounts_for_row(ws, row, rate_lav, rate_sab, rate_domfer)

        ws.cell(row, 13).value = amt_weekday   # M: $ L a V
        ws.cell(row, 14).value = amt_sat       # N: $ Sábado
        ws.cell(row, 15).value = amt_domfer    # O: $ Dom y Fer

        for c in (13, 14, 15):
            ws.cell(row, c).number_format = '"$"#,##0'
            ws.cell(row, c).alignment = Alignment(horizontal="center", vertical="center")


def update_or_build_output_workbook(weeks: dict, out_path: str, template_path: str, cfg: dict):
    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        tpl = ensure_hidden_template(wb, template_path)
    else:
        wb = load_workbook(template_path)
        tpl0 = wb.active
        tpl0.title = TEMPLATE_SHEET_NAME
        tpl = ensure_hidden_template(wb, template_path)

    for week_start, df_week in sorted(weeks.items(), key=lambda x: x[0]):
        ws, _created_new = get_or_create_week_sheet(wb, tpl, week_start, cfg)
        upsert_week_employees(ws, df_week, cfg)

    if TEMPLATE_SHEET_NAME in wb.sheetnames:
        wb[TEMPLATE_SHEET_NAME].sheet_state = "veryHidden"

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    wb.save(out_path)


# ============================================================
# UI + CALENDARIO INLINE
# ============================================================
def run_app():

    cfg = load_config()

    try:
        import ttkbootstrap as tb
    except Exception:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Falta ttkbootstrap", "Instalá: pip install ttkbootstrap")
        return

    THEME = "minty"
    root = tb.Window(themename=THEME)
    root.title("Liquidación de Horas Extra · TALCA")
    root.geometry("1120x700")
    root.minsize(1040, 660)

    FONT_TITLE = ("Segoe UI", 20, "bold")
    FONT_SUB = ("Segoe UI", 10)
    FONT_H2 = ("Segoe UI", 12, "bold")
    FONT_B = ("Segoe UI", 10, "bold")

    var_report = tk.StringVar(master=root, value="")
    var_emp = tk.StringVar(master=root, value=cfg.get("employees_excel_path", ""))
    var_tpl = tk.StringVar(master=root, value=cfg.get("template_excel_path", ""))

    var_use_holidays = tk.BooleanVar(master=root, value=False)
    var_auto_holidays = tk.BooleanVar(master=root, value=bool(cfg.get("auto_holidays_enabled", True)))

    status_var = tk.StringVar(master=root, value="Paso 1: elegí el reporte de VeoTime.")
    summary_range_var = tk.StringVar(master=root, value="—")
    summary_emps_var = tk.StringVar(master=root, value="—")
    summary_holidays_var = tk.StringVar(master=root, value="—")

    selected_holidays_manual = set()
    selected_holidays_auto = set()

    btn_generate = None

    def set_status(msg: str):
        status_var.set(msg)
        try:
            root.update_idletasks()
        except Exception:
            pass

    def month_name_es(m):
        names = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        return names[m-1]

    def get_report_date_range_and_emps(report_path: str):
        try:
            df_raw = read_report_raw(report_path)
            header_row = find_header_row(df_raw)
            if header_row is not None:
                df = apply_header_row(df_raw, header_row)
            else:
                df = df_raw.copy()
                df.columns = [str(c).strip() for c in df.iloc[0].tolist()]
                df = df.iloc[1:].copy()

            cols_norm = [normalize_text(c) for c in df.columns]
            i_fecha = guess_col(cols_norm, ["fecha"])
            i_id = guess_col(cols_norm, ["dni", "documento", "legajo", "id"])

            if i_fecha is None:
                return None, None, None

            fechas = df.iloc[:, i_fecha].apply(parse_date).dropna()
            if fechas.empty:
                return None, None, None

            start = min(fechas)
            end = max(fechas)

            emps = None
            if i_id is not None:
                tmp = df.iloc[:, i_id].apply(clean_id)
                keys = tmp.apply(id_key_from_any)
                keys = keys[keys != ""]
                if not keys.empty:
                    emps = int(keys.nunique())

            return start, end, emps
        except Exception:
            return None, None, None

    def refresh_auto_holidays():
        selected_holidays_auto.clear()

        if not var_use_holidays.get():
            return
        if not var_auto_holidays.get():
            return

        rp = var_report.get().strip()
        if not rp or not os.path.exists(rp):
            return

        start, end, _ = get_report_date_range_and_emps(rp)
        if start is None or end is None:
            return

        try:
            auto_dates = compute_auto_holidays_for_range(start, end, cfg)
            for d in auto_dates:
                selected_holidays_auto.add(d)
        except Exception:
            pass

    def update_generate_state():
        rp = var_report.get().strip()
        emp = var_emp.get().strip()
        tpl = var_tpl.get().strip()

        ok = bool(
            rp and os.path.exists(rp) and
            emp and os.path.exists(emp) and
            tpl and os.path.exists(tpl)
        )

        if btn_generate is None:
            return

        btn_generate.configure(state=("normal" if ok else "disabled"))

    def update_summary():
        rp = var_report.get().strip()
        if rp and os.path.exists(rp):
            start, end, emps = get_report_date_range_and_emps(rp)
            summary_range_var.set(f"{start.strftime('%d/%m/%Y')} → {end.strftime('%d/%m/%Y')}" if start and end else "—")
            summary_emps_var.set(str(emps) if emps is not None else "—")
        else:
            summary_range_var.set("—")
            summary_emps_var.set("—")

        all_h = sorted(selected_holidays_auto | selected_holidays_manual) if var_use_holidays.get() else []
        summary_holidays_var.set(", ".join(d.strftime("%d/%m") for d in all_h) if all_h else "—")

        update_generate_state()

    # defaults locales
    local_emp = os.path.join(os.getcwd(), EMP_MASTER_DEFAULT_NAME)
    if not var_emp.get().strip() and os.path.exists(local_emp):
        var_emp.set(local_emp)
        cfg["employees_excel_path"] = local_emp

    local_tpl = os.path.join(os.getcwd(), TEMPLATE_DEFAULT_NAME)
    if not var_tpl.get().strip() and os.path.exists(local_tpl):
        var_tpl.set(local_tpl)
        cfg["template_excel_path"] = local_tpl

    # ==========================================================
    # LAYOUT
    # ==========================================================
    header = tb.Frame(root, padding=(20, 18))
    header.pack(fill="x")

    left_h = tb.Frame(header)
    left_h.pack(side="left", fill="x", expand=True)

    tb.Label(left_h, text="Liquidación de Horas Extra", font=FONT_TITLE).pack(anchor="w")
    tb.Label(left_h, text="Generación automática desde VeoTime → Excel RRHH", font=FONT_SUB, foreground="#666").pack(anchor="w", pady=(4, 0))

    right_h = tb.Frame(header)
    right_h.pack(side="right")

    LOGO_PATH = Path(__file__).resolve().parent / "assets" / "talca_logo.png"
    logo_img = None
    try:
        from PIL import Image, ImageTk
        img = Image.open(LOGO_PATH)
        img = img.resize((160, 46), Image.LANCZOS)
        logo_img = ImageTk.PhotoImage(img)
    except Exception:
        try:
            logo_img = tk.PhotoImage(file=str(LOGO_PATH))
        except Exception:
            logo_img = None

    if logo_img:
        lbl_logo = tb.Label(right_h, image=logo_img)
        lbl_logo.pack(side="right")
        lbl_logo.image = logo_img
    else:
        tb.Label(right_h, text="TALCA", bootstyle="secondary-inverse", padding=(10, 4)).pack(side="right")

    body = tb.Frame(root, padding=(20, 0, 20, 12))
    body.pack(fill="both", expand=True)

    body.columnconfigure(0, weight=3)
    body.columnconfigure(1, weight=2)
    body.rowconfigure(0, weight=1)

    left = tb.Frame(body)
    left.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
    right = tb.Frame(body)
    right.grid(row=0, column=1, sticky="nsew", padx=(14, 0))

    nb = tb.Notebook(left, bootstyle="primary")
    nb.pack(fill="both", expand=True)

    tab_files = tb.Frame(nb, padding=16)
    tab_holidays = tb.Frame(nb, padding=16)

    nb.add(tab_files, text="  1 · Archivos  ")
    nb.add(tab_holidays, text="  2 · Feriados  ")

    summary = tb.Labelframe(right, text="RESUMEN", padding=16, bootstyle="secondary")
    summary.pack(fill="x")

    def summary_row(parent, label, var):
        fr = tb.Frame(parent)
        fr.pack(fill="x", pady=8)
        tb.Label(fr, text=label, foreground="#666").pack(side="left")
        tb.Label(fr, textvariable=var, font=FONT_B).pack(side="right")

    summary_row(summary, "Rango detectado", summary_range_var)
    summary_row(summary, "Empleados (aprox.)", summary_emps_var)
    summary_row(summary, "Feriados (dd/mm)", summary_holidays_var)

    status_card = tb.Labelframe(right, text="ESTADO", padding=16, bootstyle="light")
    status_card.pack(fill="both", expand=True, pady=(14, 0))

    tb.Label(status_card, textvariable=status_var, wraplength=360, justify="left").pack(anchor="w")
    pb = tb.Progressbar(status_card, mode="indeterminate", bootstyle="success-striped")
    pb.pack(fill="x", pady=(14, 0))

    # ==========================================================
    # TAB 1: FILES
    # ==========================================================
    tb.Label(tab_files, text="Seleccioná los 3 archivos", font=FONT_H2).pack(anchor="w")

    def entry_file(parent, title, var, browse_cmd, hint=""):
        box = tb.Frame(parent)
        box.pack(fill="x", pady=10)

        top = tb.Frame(box)
        top.pack(fill="x")
        tb.Label(top, text=title, font=FONT_B).pack(side="left")
        if hint:
            tb.Label(top, text=hint, font=FONT_SUB, foreground="#666").pack(side="left", padx=(10, 0))

        row = tb.Frame(box)
        row.pack(fill="x", pady=(6, 0))
        e = tb.Entry(row, textvariable=var)
        e.pack(side="left", fill="x", expand=True)
        tb.Button(row, text="Buscar…", command=browse_cmd, bootstyle="secondary", width=12).pack(side="left", padx=(10, 0))
        return e

    def pick_report():
        p = filedialog.askopenfilename(
            title="Seleccionar reporte de VeoTime",
            filetypes=[("Excel", "*.xls *.xlsx")]
        )
        if p:
            var_report.set(p)
            set_status("Reporte seleccionado. Paso 2: revisá feriados (si aplica).")
            refresh_auto_holidays()
            update_summary()
            try:
                render_calendar()
                render_holiday_list()
            except Exception:
                pass

    def pick_emp():
        p = filedialog.askopenfilename(
            title="Seleccionar datos empleados.xlsx",
            filetypes=[("Excel", "*.xlsx *.xls")]
        )
        if p:
            var_emp.set(p)
            cfg["employees_excel_path"] = p
            save_config(cfg)
            update_summary()

    def pick_tpl():
        p = filedialog.askopenfilename(
            title="Seleccionar formatosugerido.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if p:
            var_tpl.set(p)
            cfg["template_excel_path"] = p
            save_config(cfg)
            update_summary()

    entry_file(tab_files, "Reporte VeoTime (.xls/.xlsx)", var_report, pick_report, hint="Obligatorio")
    entry_file(tab_files, "Datos empleados.xlsx", var_emp, pick_emp, hint="Obligatorio")
    entry_file(tab_files, "Plantilla formatosugerido.xlsx", var_tpl, pick_tpl, hint="Obligatorio")

    # ==========================================================
    # TAB 2: HOLIDAYS (INLINE CALENDAR)
    # ==========================================================
    tb.Label(tab_holidays, text="Feriados de la semana", font=FONT_H2).pack(anchor="w")

    hol_top = tb.Frame(tab_holidays)
    hol_top.pack(fill="x", pady=(12, 10))

    cal_year = tk.IntVar(master=root, value=datetime.now().year)
    cal_month = tk.IntVar(master=root, value=datetime.now().month)

    def get_report_range():
        rp = var_report.get().strip()
        if rp and os.path.exists(rp):
            start, end, _ = get_report_date_range_and_emps(rp)
            return start, end
        return None, None

    def toggle_manual_date(dte: date):
        if not var_use_holidays.get():
            return
        if dte in selected_holidays_manual:
            selected_holidays_manual.remove(dte)
        else:
            selected_holidays_manual.add(dte)
        render_holiday_list()
        update_summary()
        render_calendar()

    def prev_month():
        y, m = cal_year.get(), cal_month.get()
        if m == 1:
            cal_year.set(y - 1)
            cal_month.set(12)
        else:
            cal_month.set(m - 1)
        render_calendar()

    def next_month():
        y, m = cal_year.get(), cal_month.get()
        if m == 12:
            cal_year.set(y + 1)
            cal_month.set(1)
        else:
            cal_month.set(m + 1)
        render_calendar()

    def goto_today():
        today = date.today()
        cal_year.set(today.year)
        cal_month.set(today.month)
        render_calendar()

    def on_toggle_holidays():
        sync_holiday_state()
        refresh_auto_holidays()
        render_holiday_list()
        update_summary()
        render_calendar()

    def on_toggle_auto():
        cfg["auto_holidays_enabled"] = bool(var_auto_holidays.get())
        save_config(cfg)
        refresh_auto_holidays()
        render_holiday_list()
        update_summary()
        render_calendar()

    chk_use = tb.Checkbutton(
        hol_top,
        text="Hubo feriados",
        variable=var_use_holidays,
        command=on_toggle_holidays,
        bootstyle="round-toggle"
    )
    chk_use.pack(side="left")

    chk_auto = tb.Checkbutton(
        hol_top,
        text="Detectar automáticamente (Argentina · Mendoza)",
        variable=var_auto_holidays,
        command=on_toggle_auto,
        bootstyle="round-toggle"
    )
    chk_auto.pack(side="left", padx=(14, 0))

    wrap = tb.Frame(tab_holidays)
    wrap.pack(fill="both", expand=True)

    wrap.columnconfigure(0, weight=3)
    wrap.columnconfigure(1, weight=2)
    wrap.rowconfigure(0, weight=1)

    cal_card = tb.Labelframe(wrap, text="CALENDARIO", padding=12, bootstyle="light")
    cal_card.grid(row=0, column=0, sticky="nsew", padx=(0, 12))

    list_card = tb.Labelframe(wrap, text="FERIADOS APLICADOS", padding=12, bootstyle="secondary")
    list_card.grid(row=0, column=1, sticky="nsew", padx=(12, 0))

    cal_header = tb.Frame(cal_card)
    cal_header.pack(fill="x", pady=(0, 10))

    tb.Button(cal_header, text="◀", command=prev_month, bootstyle="secondary", width=4).pack(side="left")
    lbl_month = tb.Label(cal_header, text="", font=("Segoe UI", 11, "bold"))
    lbl_month.pack(side="left", padx=10)
    tb.Button(cal_header, text="▶", command=next_month, bootstyle="secondary", width=4).pack(side="left")
    tb.Button(cal_header, text="Hoy", command=goto_today, bootstyle="info", width=8).pack(side="right")

    legend = tb.Frame(cal_card)
    legend.pack(fill="x", pady=(0, 8))
    tb.Label(legend, text="MANUAL", bootstyle="primary-inverse", padding=(8, 3)).pack(side="left")
    tb.Label(legend, text="AUTO", bootstyle="info-inverse", padding=(8, 3)).pack(side="left", padx=(8, 0))
    tb.Label(legend, text="Click para agregar/quitar", foreground="#666").pack(side="left", padx=(10, 0))

    range_lbl = tb.Label(cal_card, text="", foreground="#666")
    range_lbl.pack(anchor="w", pady=(0, 8))

    grid = tb.Frame(cal_card)
    grid.pack(fill="both", expand=True)

    days_row = tb.Frame(grid)
    days_row.pack(fill="x")
    for dn in ["Lu", "Ma", "Mi", "Ju", "Vi", "Sa", "Do"]:
        tb.Label(days_row, text=dn, width=5, anchor="center", foreground="#666").pack(side="left", padx=2)

    cells = tb.Frame(grid)
    cells.pack(fill="both", expand=True, pady=(6, 0))

    def render_calendar():
        for w in cells.winfo_children():
            w.destroy()

        y, m = cal_year.get(), cal_month.get()
        lbl_month.config(text=f"{month_name_es(m)} {y}")

        start, end = get_report_range()
        if start and end:
            range_lbl.config(text=f"Rango del reporte: {start.strftime('%d/%m/%Y')} → {end.strftime('%d/%m/%Y')}")
        else:
            range_lbl.config(text="(Tip: al elegir el reporte, se limita el rango del calendario)")

        enabled = var_use_holidays.get()
        weeks_ = calendar.monthcalendar(y, m)

        for week in weeks_:
            row = tb.Frame(cells)
            row.pack(fill="x", pady=2)

            for day in week:
                if day == 0:
                    tb.Label(row, text="", width=5).pack(side="left", padx=2)
                    continue

                dte = date(y, m, day)

                in_range = True
                if start and end:
                    in_range = (start <= dte <= end)

                is_manual = dte in selected_holidays_manual
                is_auto = dte in selected_holidays_auto

                if is_manual:
                    style = "primary"
                elif is_auto:
                    style = "info-outline"
                else:
                    style = "light"

                state = "normal"
                if (not enabled) or (not in_range):
                    state = "disabled"

                b = tb.Button(
                    row,
                    text=str(day),
                    width=5,
                    command=lambda dd=dte: toggle_manual_date(dd),
                    bootstyle=style
                )
                b.pack(side="left", padx=2)
                b.configure(state=state)

    def render_holiday_list():
        lb.delete(0, "end")
        all_dates = sorted(selected_holidays_auto | selected_holidays_manual)
        if not all_dates:
            lb.insert("end", "—")
            return
        for dte in all_dates:
            tags = []
            if dte in selected_holidays_auto:
                tags.append("AUTO")
            if dte in selected_holidays_manual:
                tags.append("MANUAL")
            suffix = f"   ·   {' + '.join(tags)}" if tags else ""
            lb.insert("end", dte.strftime("%d/%m/%Y") + suffix)

    def remove_selected_holiday():
        sel = lb.curselection()
        if not sel:
            return
        all_dates = sorted(selected_holidays_auto | selected_holidays_manual)
        if not all_dates:
            return
        dte = all_dates[sel[0]]

        if dte in selected_holidays_manual:
            selected_holidays_manual.remove(dte)
        elif dte in selected_holidays_auto:
            selected_holidays_auto.remove(dte)

        render_holiday_list()
        update_summary()
        render_calendar()

    def clear_all_holidays():
        selected_holidays_manual.clear()
        selected_holidays_auto.clear()
        render_holiday_list()
        update_summary()
        render_calendar()

    tb.Label(list_card, text="Seleccionados (auto + manual):", font=FONT_B).pack(anchor="w")
    lb = tk.Listbox(list_card, height=12)
    lb.pack(fill="both", expand=True, pady=(10, 10))

    actions = tb.Frame(list_card)
    actions.pack(fill="x")

    btn_remove_h = tb.Button(actions, text="Quitar seleccionado", command=remove_selected_holiday, bootstyle="secondary")
    btn_clear_h = tb.Button(actions, text="Limpiar todo", command=clear_all_holidays, bootstyle="secondary")
    btn_remove_h.pack(side="left")
    btn_clear_h.pack(side="right")

    def sync_holiday_state():
        enabled = var_use_holidays.get()
        st = "normal" if enabled else "disabled"

        try:
            chk_auto.configure(state=st)
        except Exception:
            pass

        try:
            btn_remove_h.configure(state=st)
            btn_clear_h.configure(state=st)
        except Exception:
            pass

        try:
            lb.configure(state=st)
        except Exception:
            pass

        if not enabled:
            selected_holidays_manual.clear()
            selected_holidays_auto.clear()

        render_holiday_list()
        update_summary()
        render_calendar()

    # ==========================================================
    # FOOTER: GENERATE
    # ==========================================================
    footer = tb.Frame(root, padding=(20, 12))
    footer.pack(fill="x")

    tb.Label(footer, text="Podés guardar un archivo nuevo o elegir uno existente para agregar empleados.", foreground="#666").pack(side="left")

    def generate():
        rp = var_report.get().strip()
        emp = var_emp.get().strip()
        tpl = var_tpl.get().strip()

        default_out = os.path.join(
            os.path.dirname(rp) if rp else os.getcwd(),
            "Liquidacion_Horas_Extra.xlsx"
        )

        out_path = filedialog.asksaveasfilename(
            title="Guardar / Actualizar liquidación",
            defaultextension=".xlsx",
            initialfile=os.path.basename(default_out),
            filetypes=[("Excel", "*.xlsx")]
        )
        if not out_path:
            return

        existed = os.path.exists(out_path)

        btn_generate.configure(state="disabled")
        pb.start(12)

        try:
            set_status("Procesando… leyendo empleados y reporte VeoTime…")

            if var_use_holidays.get():
                refresh_auto_holidays()
                all_h = sorted(selected_holidays_auto | selected_holidays_manual)
                cfg["holidays"] = [d.strftime("%Y-%m-%d") for d in all_h]
            else:
                cfg["holidays"] = []

            cfg["auto_holidays_enabled"] = bool(var_auto_holidays.get())
            save_config(cfg)

            emp_master = load_employee_master(emp)
            daily = read_veotime_to_daily(rp, emp_master, cfg)
            weeks = compute_overtime_from_daily(daily, cfg)

            set_status("Escribiendo Excel… (si ya existe, agrega empleados abajo)")
            update_or_build_output_workbook(weeks, out_path, tpl, cfg)

            set_status("Listo ✅ Excel generado/actualizado correctamente.")
            messagebox.showinfo(
                "Listo ✅",
                ("Se ACTUALIZÓ el Excel (se agregaron empleados debajo de los existentes):\n"
                 if existed else "Se generó el Excel:\n") + f"{out_path}"
            )

        except PermissionError:
            set_status("Error ❌ El archivo está abierto.")
            messagebox.showerror(
                "Archivo en uso",
                "Cerrá el Excel (está abierto) y volvé a intentar.\n"
                "Windows no deja guardar si el archivo está en uso."
            )

        except Exception as e:
            tbtxt = traceback.format_exc()
            with open("debug_error.txt", "w", encoding="utf-8") as f:
                f.write(tbtxt)
            set_status("Error ❌ Revisá debug_error.txt")
            messagebox.showerror("Error", f"{e}\n\nSe guardó el detalle en debug_error.txt")

        finally:
            pb.stop()
            update_generate_state()

    btn_generate = tb.Button(footer, text="Generar / Actualizar", command=generate, bootstyle="success", width=22)
    btn_generate.pack(side="right")

    # ==========================================================
    # Bindings + init
    # ==========================================================
    def on_any_change(*_):
        update_summary()

    var_report.trace_add("write", on_any_change)
    var_emp.trace_add("write", on_any_change)
    var_tpl.trace_add("write", on_any_change)

    refresh_auto_holidays()
    render_holiday_list()
    render_calendar()
    sync_holiday_state()
    update_summary()

    root.mainloop()


if __name__ == "__main__":
    run_app()
