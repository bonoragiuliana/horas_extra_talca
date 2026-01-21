import os
import re
from copy import copy
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.formula.translate import Translator

from config_app import TEMPLATE_SHEET_NAME
from holidays_auto import is_holiday
from utils_app import normalize_text, excel_cell_to_date, id_key_from_any, num_or_zero


# ============================================================
# TEMPLATE HELPERS
# ============================================================
def _safe_sheet_title(s: str) -> str:
    s = re.sub(r'[\\/*?:\[\]]', '-', str(s))
    return s[:31].strip() or "Semana"


def get_week_holidays(week_start: date, cfg: dict):
    """
    Devuelve feriados de la semana (Lun..Sab).
    Ojo: Dom ya tiene su columna propia, así que no creamos columna "FERIADO" para domingo.
    """
    hols = []
    for i in range(0, 6):  # Lun..Sab
        d = week_start + timedelta(days=i)
        if is_holiday(d, cfg):
            hols.append(d)
    hols.sort()
    return hols


def get_holiday_cols(ws, base_col=12, header_row=3):
    """
    Devuelve las columnas consecutivas que dicen 'FERIADO' en el header (fila 3),
    empezando en la columna 12.
    """
    cols = []
    c = base_col
    while True:
        v = normalize_text(ws.cell(header_row, c).value)
        if v == "feriado":
            cols.append(c)
            c += 1
        else:
            break

    if not cols:
        cols = [base_col]
    return cols


def _col_for_date(ws, d: date, cfg: dict) -> int:
    # Dom siempre va a su columna (aunque sea feriado)
    if d.weekday() == 6:
        return 5

    # Si es feriado (Lun..Sab), buscar su columna por fecha en el header
    if is_holiday(d, cfg):
        hol_cols = get_holiday_cols(ws, base_col=12)
        for c in hol_cols:
            hd = excel_cell_to_date(ws.cell(4, c).value)
            if hd == d:
                return c
        return hol_cols[0] if hol_cols else 12

    # Lun..Sab normal
    return 6 + d.weekday()  # Lun=6 ... Sab=11


def _ensure_rows(ws, start_row, n_rows_needed, base_style_row=7):
    max_col = max(ws.max_column, 30)
    last_needed = start_row + n_rows_needed - 1
    if last_needed <= ws.max_row:
        return

    for r in range(ws.max_row + 1, last_needed + 1):
        ws.insert_rows(r)
        for c in range(1, max_col + 1):
            src = ws.cell(base_style_row, c)
            dst = ws.cell(r, c)

            dst._style = copy(src._style)
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)

            v = src.value
            if isinstance(v, str) and v.startswith("="):
                try:
                    dst.value = Translator(v, origin=src.coordinate).translate_formula(dst.coordinate)
                except Exception:
                    dst.value = v
            else:
                dst.value = v


def _set_week_header(ws, week_start: date, cfg: dict):
    # Dom..Sab
    sun = week_start - timedelta(days=1)
    dates = [
        (5, sun),
        (6, week_start),
        (7, week_start + timedelta(days=1)),
        (8, week_start + timedelta(days=2)),
        (9, week_start + timedelta(days=3)),
        (10, week_start + timedelta(days=4)),
        (11, week_start + timedelta(days=5)),
    ]
    for col, d in dates:
        cell = ws.cell(4, col)
        cell.value = d
        cell.number_format = "dd/mm/yy"

    hol_dates = get_week_holidays(week_start, cfg)
    desired = max(1, len(hol_dates))

    hol_cols = get_holiday_cols(ws, base_col=12)
    current = len(hol_cols)

    if desired > current:
        insert_at = 12 + current
        for _ in range(desired - current):
            ws.insert_cols(insert_at)

            for r in range(1, ws.max_row + 1):
                src = ws.cell(r, 12)
                dst = ws.cell(r, insert_at)
                if src.has_style:
                    dst._style = copy(src._style)
                    dst.font = copy(src.font)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                    dst.number_format = src.number_format
                    dst.protection = copy(src.protection)
                    dst.alignment = copy(src.alignment)

            if not ws.cell(3, insert_at).value:
                ws.cell(3, insert_at).value = ws.cell(3, 12).value or "FERIADO"
            if not ws.cell(5, insert_at).value:
                ws.cell(5, insert_at).value = ws.cell(5, 12).value or "Hs"

            insert_at += 1

        hol_cols = get_holiday_cols(ws, base_col=12)

    for i, c in enumerate(hol_cols):
        d = hol_dates[i] if i < len(hol_dates) else None
        cell = ws.cell(4, c)
        cell.value = d
        cell.number_format = "dd/mm/yy"


def _clear_data_area(ws, start_row=7):
    hol_cols = get_holiday_cols(ws, base_col=12)
    last_hol = hol_cols[-1] if hol_cols else 12

    col_amt_wd = last_hol + 1
    col_amt_sat = last_hol + 2
    col_amt_domfer = last_hol + 3
    col_subtotal = last_hol + 4
    col_redondeo = last_hol + 5

    for r in range(start_row, ws.max_row + 1):
        for c in range(1, 5):
            ws.cell(r, c).value = None

        for c in range(5, 12):
            ws.cell(r, c).value = 0
            ws.cell(r, c).number_format = "0"
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")

        for c in hol_cols:
            ws.cell(r, c).value = 0
            ws.cell(r, c).number_format = "0"
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")

        for c in (col_amt_wd, col_amt_sat, col_amt_domfer, col_subtotal, col_redondeo):
            ws.cell(r, c).value = 0
            ws.cell(r, c).number_format = '"$"#,##0'
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")


# ============================================================
# TEMPLATE: asegurar _TEMPLATE oculta dentro del workbook de salida
# ============================================================
def clone_template_sheet_into_workbook(wb_dest, template_path: str) -> None:
    src_wb = load_workbook(template_path)
    src = src_wb.active

    if TEMPLATE_SHEET_NAME in wb_dest.sheetnames:
        src_wb.close()
        return

    dst = wb_dest.create_sheet(TEMPLATE_SHEET_NAME)

    for col_letter, dim in src.column_dimensions.items():
        dst.column_dimensions[col_letter].width = dim.width

    for row_idx, dim in src.row_dimensions.items():
        dst.row_dimensions[row_idx].height = dim.height

    for row in src.iter_rows():
        for cell in row:
            new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    for m in src.merged_cells.ranges:
        dst.merge_cells(str(m))

    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    dst.page_setup = copy(src.page_setup)
    dst.page_margins = copy(src.page_margins)

    src_wb.close()


def ensure_hidden_template(wb, template_path: str):
    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        if "TEMPLATE" in wb.sheetnames:
            ws = wb["TEMPLATE"]
            ws.title = TEMPLATE_SHEET_NAME

    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        clone_template_sheet_into_workbook(wb, template_path)

    tpl = wb[TEMPLATE_SHEET_NAME]
    tpl.sheet_state = "veryHidden"
    return tpl


def find_week_sheet(wb, week_start: date):
    target_title = _safe_sheet_title(f"Semana {week_start.strftime('%d-%m')}")
    if target_title in wb.sheetnames:
        return wb[target_title]

    for name in wb.sheetnames:
        if name == TEMPLATE_SHEET_NAME:
            continue
        ws = wb[name]
        v = ws.cell(4, 6).value  # F4 = lunes
        d = excel_cell_to_date(v)
        if d == week_start:
            return ws
    return None


def get_or_create_week_sheet(wb, tpl, week_start: date, cfg: dict):
    ws = find_week_sheet(wb, week_start)
    created_new = False

    if ws is None:
        ws = wb.copy_worksheet(tpl)
        ws.sheet_state = "visible"
        ws.title = _safe_sheet_title(f"Semana {week_start.strftime('%d-%m')}")
        _set_week_header(ws, week_start, cfg)
        _clear_data_area(ws, start_row=7)
        created_new = True
    else:
        _set_week_header(ws, week_start, cfg)

    return ws, created_new


def read_existing_employee_rows(ws, start_row=7):
    row_by_key = {}
    last_filled = start_row - 1
    empty_run = 0

    for r in range(start_row, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value

        has_data = any(v not in (None, "") for v in [a, b, c, d])

        if has_data:
            empty_run = 0
            last_filled = r
            k = id_key_from_any(c)
            if k:
                row_by_key[k] = r
        else:
            empty_run += 1
            if empty_run >= 30 and r > start_row + 30:
                break

    return row_by_key, (last_filled + 1)


def _ensure_row_formulas_from_base(ws, target_row, base_row=7, max_col=30, skip_cols=None):
    skip_cols = set(skip_cols or [])
    for c in range(1, max_col + 1):
        if c in skip_cols:
            continue
        src = ws.cell(base_row, c)
        v = src.value
        if isinstance(v, str) and v.startswith("="):
            dst = ws.cell(target_row, c)
            if isinstance(dst.value, str) and dst.value.startswith("="):
                continue
            try:
                dst.value = Translator(v, origin=src.coordinate).translate_formula(dst.coordinate)
            except Exception:
                dst.value = v


def force_integer_hours_format(ws, row, col_start=5, col_end=12):
    """
    Fuerza que las celdas de horas se vean como enteros (sin ,00).
    """
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row, c)
        try:
            if cell.value not in (None, ""):
                cell.value = int(float(cell.value))
        except Exception:
            pass

        cell.number_format = "0"
        cell.alignment = Alignment(horizontal="center", vertical="center")


def compute_amounts_for_row(ws, row, rate_lav, rate_sab, rate_domfer, holiday_cols):
    # Dom=5, Lun=6, Mar=7, Mier=8, Jue=9, Vie=10, Sab=11
    hours_dom = int(num_or_zero(ws.cell(row, 5).value))
    hours_lun = int(num_or_zero(ws.cell(row, 6).value))
    hours_mar = int(num_or_zero(ws.cell(row, 7).value))
    hours_mie = int(num_or_zero(ws.cell(row, 8).value))
    hours_jue = int(num_or_zero(ws.cell(row, 9).value))
    hours_vie = int(num_or_zero(ws.cell(row, 10).value))
    hours_sab = int(num_or_zero(ws.cell(row, 11).value))

    hours_fer_total = 0
    for c in holiday_cols:
        hours_fer_total += int(num_or_zero(ws.cell(row, c).value))

    hours_weekday = hours_lun + hours_mar + hours_mie + hours_jue + hours_vie
    hours_domfer = hours_dom + hours_fer_total

    amt_weekday = int(round(hours_weekday * float(rate_lav), 0))
    amt_sat = int(round(hours_sab * float(rate_sab), 0))
    amt_domfer = int(round(hours_domfer * float(rate_domfer), 0))

    return amt_weekday, amt_sat, amt_domfer


def upsert_week_employees(ws, df_week: pd.DataFrame, cfg: dict):
    if df_week is None or df_week.empty:
        return

    holiday_cols = get_holiday_cols(ws, base_col=12)
    last_hol = holiday_cols[-1] if holiday_cols else 12

    col_amt_wd = last_hol + 1
    col_amt_sat = last_hol + 2
    col_amt_domfer = last_hol + 3
    col_subtotal = last_hol + 4
    col_redondeo = last_hol + 5

    sub = df_week.copy()

    day_he = (
        sub.groupby(
            ["empresa", "dni", "dni_display", "nombre", "sector", "fecha", "rate_lav", "rate_sab", "rate_domfer"],
            dropna=False
        )["horas_extra"]
        .sum()
        .reset_index()
    )

    employees = (
        day_he[["empresa", "dni", "dni_display", "nombre", "sector", "rate_lav", "rate_sab", "rate_domfer"]]
        .drop_duplicates()
        .sort_values(["empresa", "sector", "nombre"])
    )

    he_map = {}
    for _, r in day_he.iterrows():
        k = str(r["dni"]).strip()
        he_map.setdefault(k, {})[r["fecha"]] = int(r["horas_extra"])

    row_by_key, next_row = read_existing_employee_rows(ws, start_row=7)

    _ensure_rows(
        ws,
        start_row=7,
        n_rows_needed=max(10, next_row - 6 + len(employees)),
        base_style_row=7
    )

    merge_mode = str(cfg.get("merge_mode", "sum")).strip().lower()

    for _, emp in employees.iterrows():
        empresa = str(emp["empresa"])
        key_internal = str(emp["dni"]).strip()
        dni_display = str(emp.get("dni_display", "")).strip() or key_internal
        nombre = str(emp["nombre"])
        sector = str(emp["sector"])

        rate_lav = float(emp["rate_lav"])
        rate_sab = float(emp["rate_sab"])
        rate_domfer = float(emp["rate_domfer"])

        if not key_internal:
            continue

        if key_internal in row_by_key:
            row = row_by_key[key_internal]
            is_new = False
        else:
            row = next_row
            next_row += 1
            row_by_key[key_internal] = row
            is_new = True

        ws.cell(row, 1).value = empresa
        ws.cell(row, 2).value = sector
        ws.cell(row, 3).value = dni_display
        ws.cell(row, 4).value = nombre

        if is_new:
            for c in range(5, 12):
                ws.cell(row, c).value = 0
                ws.cell(row, c).number_format = "0"
                ws.cell(row, c).alignment = Alignment(horizontal="center", vertical="center")
            for c in holiday_cols:
                ws.cell(row, c).value = 0
                ws.cell(row, c).number_format = "0"
                ws.cell(row, c).alignment = Alignment(horizontal="center", vertical="center")

        emp_he_by_date = he_map.get(key_internal, {})
        for dte, he in emp_he_by_date.items():
            he = int(he)
            if he <= 0:
                continue

            col = _col_for_date(ws, dte, cfg)
            cur = int(num_or_zero(ws.cell(row, col).value))

            if merge_mode == "replace":
                new_val = he
            else:
                new_val = cur + he

            ws.cell(row, col).value = int(new_val)
            ws.cell(row, col).number_format = "0"
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")

        force_integer_hours_format(ws, row, 5, last_hol)

        _ensure_row_formulas_from_base(
            ws,
            target_row=row,
            base_row=7,
            max_col=max(ws.max_column, col_redondeo, 30),
            skip_cols={col_amt_wd, col_amt_sat, col_amt_domfer, col_subtotal, col_redondeo}
        )

        amt_weekday, amt_sat, amt_domfer = compute_amounts_for_row(
            ws, row, rate_lav, rate_sab, rate_domfer, holiday_cols
        )
        subtotal = int(amt_weekday + amt_sat + amt_domfer)
        redondeo = subtotal

        ws.cell(row, col_amt_wd).value = int(amt_weekday)
        ws.cell(row, col_amt_sat).value = int(amt_sat)
        ws.cell(row, col_amt_domfer).value = int(amt_domfer)
        ws.cell(row, col_subtotal).value = int(subtotal)
        ws.cell(row, col_redondeo).value = int(redondeo)

        for c in (col_amt_wd, col_amt_sat, col_amt_domfer, col_subtotal, col_redondeo):
            ws.cell(row, c).number_format = '"$"#,##0'
            ws.cell(row, c).alignment = Alignment(horizontal="center", vertical="center")


def update_or_build_output_workbook(weeks: dict, out_path: str, template_path: str, cfg: dict):
    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        tpl = ensure_hidden_template(wb, template_path)
    else:
        wb = load_workbook(template_path)
        tpl0 = wb.active
        tpl0.title = TEMPLATE_SHEET_NAME
        tpl = ensure_hidden_template(wb, template_path)

    for week_start, df_week in sorted(weeks.items(), key=lambda x: x[0]):
        ws, _created_new = get_or_create_week_sheet(wb, tpl, week_start, cfg)
        upsert_week_employees(ws, df_week, cfg)

    if TEMPLATE_SHEET_NAME in wb.sheetnames:
        wb[TEMPLATE_SHEET_NAME].sheet_state = "veryHidden"

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    wb.save(out_path)
