# rrhh_print.py
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ----------------------------
# Helpers de texto / parsing
# ----------------------------
def _norm_text(x) -> str:
    if x is None:
        return ""
    if not isinstance(x, str):
        x = str(x)
    s = x.strip().upper()
    s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s


def _to_number_ar(v) -> float:
    """Convierte valores tipo $92.500 / 166.500 / 2 / '0' a float."""
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return float(int(v))
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    if s == "" or s.startswith("#"):
        return 0.0

    s = s.replace("$", "").replace(" ", "")

    # Caso argentino: 92.500 (miles con punto)
    if "." in s and "," in s:
        # 1.234,56 -> 1234.56
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        # 12,5 -> 12.5
        s = s.replace(",", ".")
    elif "." in s and "," not in s:
        # Si matchea miles: 1.234.567 -> 1234567
        if re.fullmatch(r"-?\d{1,3}(\.\d{3})+", s):
            s = s.replace(".", "")

    s = re.sub(r"[^0-9\.\-]", "", s)
    try:
        return float(s)
    except Exception:
        return 0.0


# ----------------------------
# Detección de layout
# ----------------------------
DAY_ALIASES = {
    "Dom": ["DOM"],
    "Lun": ["LUN"],
    "Mar": ["MAR"],
    "Miérc": ["MIERC", "MIER", "MIE"],
    "Juev": ["JUEV", "JUE"],
    "Vier": ["VIER", "VIE"],
    "Sáb": ["SAB", "SÁB"],
}


def _detect_day_header_row(ws, max_row=80, max_col=120) -> Optional[int]:
    """Fila donde aparecen Dom/Lun/Mar/... (ej: fila 3)."""
    for r in range(1, min(ws.max_row, max_row) + 1):
        found = set()
        for c in range(1, min(ws.max_column, max_col) + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            t = _norm_text(v)
            for lab, aliases in DAY_ALIASES.items():
                if any(t.startswith(a) for a in aliases):
                    found.add(lab)
        if "Dom" in found and len(found) >= 5:
            return r
    return None


def _detect_name_col(ws, max_row=80, max_col=120) -> Optional[int]:
    """Columna de 'APELLIDO Y NOMBRE'."""
    for r in range(1, min(ws.max_row, max_row) + 1):
        for c in range(1, min(ws.max_column, max_col) + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            t = _norm_text(v)
            if "APELLIDO" in t and "NOMBRE" in t:
                return c
    return None


def _detect_num_col(ws, max_row=40, max_col=40) -> Optional[int]:
    """Columna de 'NUM' (en impresión)."""
    for r in range(1, min(ws.max_row, max_row) + 1):
        for c in range(1, min(ws.max_column, max_col) + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            t = _norm_text(v)
            if t == "NUM" or t.startswith("NUM"):
                return c
    return None


def _detect_day_cols(ws, day_row: int, max_col=200) -> Dict[str, int]:
    cols: Dict[str, int] = {}
    for c in range(1, min(ws.max_column, max_col) + 1):
        v = ws.cell(day_row, c).value
        if not isinstance(v, str):
            continue
        t = _norm_text(v)
        for lab, aliases in DAY_ALIASES.items():
            if any(t.startswith(a) for a in aliases):
                cols[lab] = c
    return cols


def _detect_holiday_col(ws, day_row: int, max_col=200) -> Optional[int]:
    for c in range(1, min(ws.max_column, max_col) + 1):
        v = ws.cell(day_row, c).value
        if isinstance(v, str) and "FERIADO" in _norm_text(v):
            return c
    return None


def _detect_totals_header_row(ws, start_row: int, look_ahead=4, max_col=250) -> int:
    """En impresión suele ser day_row+1 (fila 4). En semanal puede ser 3 o 4."""
    best_row = start_row
    best_score = 0
    for r in range(start_row, start_row + look_ahead):
        score = 0
        for c in range(1, min(ws.max_column, max_col) + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            t = _norm_text(v)
            if "$/H" in t:
                score += 1
            if "SUB TOTAL" in t or "SUBTOTAL" in t:
                score += 1
            if "REDONDE" in t:
                score += 1
            if t == "TOTAL":
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def _detect_totals_cols(ws, totals_row: int, max_col=250) -> Dict[str, int]:
    cols: Dict[str, int] = {}
    for c in range(1, min(ws.max_column, max_col) + 1):
        v = ws.cell(totals_row, c).value
        if not isinstance(v, str):
            continue
        t = _norm_text(v)

        if "$/H" in t and "L A V" in t:
            cols["lv"] = c
        elif "$/H" in t and "SAB" in t:
            cols["sab"] = c
        elif "$/H" in t and ("DOM" in t or "FER" in t):
            cols["domfer"] = c
        elif "SUB TOTAL" in t or t == "SUBTOTAL":
            cols["subtotal"] = c
        elif "REDONDE" in t:
            cols["redondeo"] = c
        elif t == "TOTAL":
            cols["total"] = c

    return cols


def _find_best_table_sheet(wb) -> Tuple[object, int]:
    """Elige la primer hoja que parezca una planilla semanal (tenga Dom/Lun/...)."""
    for ws in wb.worksheets:
        r = _detect_day_header_row(ws)
        if r:
            return ws, r
    raise ValueError("No encontré en el archivo una hoja con encabezados Dom/Lun/Mar... (planilla semanal).")


def _guess_data_start(day_header_row: int) -> int:
    # En tu formato: day row = 3, fechas=4, Hs=5, precio=6, datos=7
    return day_header_row + 4


# ----------------------------
# Modelo de fila de empleado
# ----------------------------
@dataclass
class EmpRow:
    name: str
    hours: Dict[str, float]  # Dom..Sáb y "Feriado"
    totals: Dict[str, float]  # lv/sab/domfer/subtotal/redondeo/total


def _name_key(name: str) -> str:
    return _norm_text(name)


# ----------------------------
# Lectura de planilla semanal
# ----------------------------
def _read_weekly_file(path_xlsx: str) -> Tuple[Dict[str, object], List[EmpRow]]:
    """
    Devuelve:
      meta: {'day_dates': {Dom:date, ...}, 'holiday_date': date|None}
      rows: [EmpRow...]
    """
    wb = load_workbook(path_xlsx, data_only=True)
    ws, day_row = _find_best_table_sheet(wb)

    name_col = _detect_name_col(ws)
    if not name_col:
        raise ValueError(f"No encontré la columna 'APELLIDO Y NOMBRE' en {path_xlsx}")

    day_cols = _detect_day_cols(ws, day_row)
    if "Dom" not in day_cols:
        raise ValueError(f"No encontré la columna 'Dom' en {path_xlsx}")

    hol_col = _detect_holiday_col(ws, day_row)  # 1 sola en tus formatos
    totals_row = _detect_totals_header_row(ws, day_row)
    totals_cols = _detect_totals_cols(ws, totals_row)

    data_start = _guess_data_start(day_row)

    # Fechas (fila debajo del day_row)
    day_dates: Dict[str, object] = {}
    for lab in ["Dom", "Lun", "Mar", "Miérc", "Juev", "Vier", "Sáb"]:
        c = day_cols.get(lab)
        if c:
            day_dates[lab] = ws.cell(day_row + 1, c).value

    holiday_date = ws.cell(day_row + 1, hol_col).value if hol_col else None

    rows: List[EmpRow] = []
    empty_streak = 0

    for r in range(data_start, ws.max_row + 1):
        name_v = ws.cell(r, name_col).value
        name = (str(name_v).strip() if isinstance(name_v, str) else "")
        if not name:
            empty_streak += 1
            if empty_streak >= 30:
                break
            continue
        empty_streak = 0

        hours = {}
        for lab in ["Dom", "Lun", "Mar", "Miérc", "Juev", "Vier", "Sáb"]:
            c = day_cols.get(lab)
            hours[lab] = _to_number_ar(ws.cell(r, c).value) if c else 0.0
        hours["Feriado"] = _to_number_ar(ws.cell(r, hol_col).value) if hol_col else 0.0

        totals = {
            "lv": _to_number_ar(ws.cell(r, totals_cols.get("lv", 0)).value) if totals_cols.get("lv") else 0.0,
            "sab": _to_number_ar(ws.cell(r, totals_cols.get("sab", 0)).value) if totals_cols.get("sab") else 0.0,
            "domfer": _to_number_ar(ws.cell(r, totals_cols.get("domfer", 0)).value) if totals_cols.get("domfer") else 0.0,
            "subtotal": _to_number_ar(ws.cell(r, totals_cols.get("subtotal", 0)).value) if totals_cols.get("subtotal") else 0.0,
            "redondeo": _to_number_ar(ws.cell(r, totals_cols.get("redondeo", 0)).value) if totals_cols.get("redondeo") else 0.0,
            "total": _to_number_ar(ws.cell(r, totals_cols.get("total", 0)).value) if totals_cols.get("total") else 0.0,
        }

        # Si no vino el total, lo calculamos
        if totals["total"] == 0.0 and (totals["subtotal"] != 0.0 or totals["redondeo"] != 0.0):
            totals["total"] = totals["subtotal"] + totals["redondeo"]

        rows.append(EmpRow(name=name, hours=hours, totals=totals))

    meta = {"day_dates": day_dates, "holiday_date": holiday_date}
    return meta, rows


# ----------------------------
# Escritura en plantilla impresión
# ----------------------------
def build_rrhh_print_workbook(
    path_oeste: str,
    path_consultora: str,
    template_print_path: str,
    out_path: str,
    template_sheet_name: str = "impresion",
    clear_extra_rows: int = 400,
) -> None:
    """
    Genera el XLSX de impresión RRHH usando tu plantilla de impresión.

    - Junta empleados de Oeste + Consultora
    - Ordena alfabéticamente
    - Copia horas + importes + totales (K..P) COMO VALORES (no fórmulas)
    """
    meta_o, rows_o = _read_weekly_file(path_oeste)
    meta_c, rows_c = _read_weekly_file(path_consultora)

    # Tomamos las fechas de la primera (asumimos misma semana)
    meta = meta_o if meta_o.get("day_dates") else meta_c

    # Merge por nombre (si llegara a repetirse, suma)
    merged: Dict[str, EmpRow] = {}

    def add_rows(rows: List[EmpRow]) -> None:
        for er in rows:
            k = _name_key(er.name)
            if k not in merged:
                merged[k] = er
            else:
                # suma horas y totales si aparece repetido
                cur = merged[k]
                for d, v in er.hours.items():
                    cur.hours[d] = _to_number_ar(cur.hours.get(d, 0.0)) + _to_number_ar(v)
                for t, v in er.totals.items():
                    cur.totals[t] = _to_number_ar(cur.totals.get(t, 0.0)) + _to_number_ar(v)

    add_rows(rows_o)
    add_rows(rows_c)

    final_rows = sorted(merged.values(), key=lambda x: _name_key(x.name))

    # Abrimos plantilla de impresión
    wb = load_workbook(template_print_path)
    ws = wb[template_sheet_name] if template_sheet_name in wb.sheetnames else wb.active

    # Detección layout plantilla
    day_row = _detect_day_header_row(ws)
    if not day_row:
        raise ValueError("La plantilla de impresión no tiene encabezados Dom/Lun/Mar...")

    name_col = _detect_name_col(ws)
    num_col = _detect_num_col(ws)

    if not name_col or not num_col:
        raise ValueError("La plantilla de impresión debe tener columnas 'NUM' y 'APELLIDO Y NOMBRE'.")

    day_cols = _detect_day_cols(ws, day_row)
    hol_col = _detect_holiday_col(ws, day_row)

    totals_row = _detect_totals_header_row(ws, day_row)
    totals_cols = _detect_totals_cols(ws, totals_row)

    data_start = _guess_data_start(day_row)

    # Poner fechas en la plantilla (fila day_row+1)
    dates_row = day_row + 1
    for lab, dt in (meta.get("day_dates") or {}).items():
        c = day_cols.get(lab)
        if c:
            ws.cell(dates_row, c).value = dt
    if hol_col:
        ws.cell(dates_row, hol_col).value = meta.get("holiday_date")

    # Columnas a limpiar / usar
    relevant_cols = {num_col, name_col}
    relevant_cols.update(day_cols.values())
    if hol_col:
        relevant_cols.add(hol_col)
    relevant_cols.update(totals_cols.values())

    # Limpieza de filas (para que no queden #VALOR! en 200 filas vacías)
    # Limpiamos desde data_start hasta data_start+clear_extra_rows, o hasta max_row actual si es mayor.
    last_clear = max(ws.max_row, data_start + clear_extra_rows)
    for r in range(data_start, last_clear + 1):
        for c in relevant_cols:
            ws.cell(r, c).value = None

    # Escritura empleados
    for idx, er in enumerate(final_rows, start=1):
        r = data_start + idx - 1
        ws.cell(r, num_col).value = idx
        ws.cell(r, name_col).value = er.name

        for lab in ["Dom", "Lun", "Mar", "Miérc", "Juev", "Vier", "Sáb"]:
            c = day_cols.get(lab)
            if c:
                ws.cell(r, c).value = _to_number_ar(er.hours.get(lab, 0.0))

        if hol_col:
            ws.cell(r, hol_col).value = _to_number_ar(er.hours.get("Feriado", 0.0))

        # Totales (copiados como valores)
        for key in ["lv", "sab", "domfer", "subtotal", "redondeo", "total"]:
            c = totals_cols.get(key)
            if c:
                ws.cell(r, c).value = _to_number_ar(er.totals.get(key, 0.0))

    # Print area: desde la primera col relevante hasta la última, hasta el último empleado
    if final_rows:
        last_row = data_start + len(final_rows) - 1
    else:
        last_row = data_start

    left = min(relevant_cols)
    right = max(relevant_cols)
    ws.print_area = f"{get_column_letter(left)}1:{get_column_letter(right)}{last_row}"

    wb.save(out_path)
