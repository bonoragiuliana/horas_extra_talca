# print_rrhh.py
from __future__ import annotations

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ==========================================================
# IMPRESIÓN RRHH: GENERAR DESDE 2 PLANILLAS YA GENERADAS
# (OESTE + CONSULTORA)
# ==========================================================

def _norm(s):
    return (s or "").strip().upper()

def _is_text(v):
    return isinstance(v, str) and v.strip() != ""


def find_table_sheet_and_header_row(wb):
    """
    Busca la hoja y la fila donde aparece 'APELLIDO Y NOMBRE'.
    Devuelve (ws, header_row, name_col).
    """
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 120) + 1):
            for c in range(1, min(ws.max_column, 120) + 1):
                v = ws.cell(r, c).value
                if _is_text(v) and "APELLIDO" in _norm(v) and "NOMBRE" in _norm(v):
                    return ws, r, c
    raise ValueError(
        "No encontré 'APELLIDO Y NOMBRE' en ninguna hoja.\n"
        "Asegurate de seleccionar una planilla GENERADA por el sistema."
    )


def _row_has_days(ws, r, max_c=160):
    """
    Heurística: si en esa fila aparece DOM y también (LUN o MAR) cerca, probablemente sea la fila de días.
    """
    texts = []
    for c in range(1, min(ws.max_column, max_c) + 1):
        v = ws.cell(r, c).value
        texts.append(_norm(v) if _is_text(v) else "")

    # Buscar una celda que arranque con DOM
    for idx, t in enumerate(texts):
        if t.startswith("DOM"):
            # chequear próximos 10
            window = texts[idx: idx + 10]
            if any(x.startswith("LUN") for x in window) or any(x.startswith("MAR") for x in window):
                return True
    return False


def find_day_and_totals_columns(ws, header_row_name):
    """
    Busca la fila real donde están los días (Dom..Sáb) y luego detecta:
    - dom_col
    - totals_start (donde empiezan $/h / Sub Total / Redondeo)
    - holiday_cols (columna/s de feriado)
    Devuelve: (day_header_row, dom_col, totals_start, holiday_cols)
    """
    max_c = min(ws.max_column, 180)

    # 1) localizar la fila de días cerca del encabezado de "APELLIDO Y NOMBRE"
    candidate_rows = list(range(max(1, header_row_name - 5), min(ws.max_row, header_row_name + 12) + 1))
    day_header_row = None
    for r in candidate_rows:
        if _row_has_days(ws, r, max_c=max_c):
            day_header_row = r
            break

    # fallback: buscar en las primeras ~100 filas
    if day_header_row is None:
        for r in range(1, min(ws.max_row, 120) + 1):
            if _row_has_days(ws, r, max_c=max_c):
                day_header_row = r
                break

    if day_header_row is None:
        raise ValueError(
            "No encontré la fila con los días (Dom/Lun/...).\n"
            "Asegurate de seleccionar una planilla GENERADA por el sistema."
        )

    # 2) encontrar columna DOM en esa fila
    dom_col = None
    for c in range(1, max_c + 1):
        v = ws.cell(day_header_row, c).value
        txt = _norm(v) if _is_text(v) else ""
        if txt.startswith("DOM"):
            dom_col = c
            break

    if dom_col is None:
        raise ValueError(
            "No encontré la columna 'Dom' en el encabezado.\n"
            "Asegurate de seleccionar una planilla GENERADA por el sistema."
        )

    # 3) encontrar inicio de totales
    totals_start = None
    for c in range(dom_col, max_c + 1):
        v = ws.cell(day_header_row, c).value
        txt = _norm(v) if _is_text(v) else ""
        if ("$/H" in txt) or ("SUB TOTAL" in txt) or ("SUBTOTAL" in txt) or ("REDONDE" in txt):
            totals_start = c
            break

    if totals_start is None:
        # A veces $/h está en la fila siguiente (por merges). Probamos 2 filas más.
        for rr in (day_header_row + 1, day_header_row + 2):
            for c in range(dom_col, max_c + 1):
                v = ws.cell(rr, c).value
                txt = _norm(v) if _is_text(v) else ""
                if ("$/H" in txt) or ("SUB TOTAL" in txt) or ("SUBTOTAL" in txt) or ("REDONDE" in txt):
                    totals_start = c
                    break
            if totals_start is not None:
                break

    if totals_start is None:
        raise ValueError(
            "No encontré dónde empiezan los totales ($/h / Sub Total / Redondeo).\n"
            "Asegurate de seleccionar una planilla GENERADA por el sistema."
        )

    # 4) feriados (si hay)
    day_block_cols = list(range(dom_col, totals_start))
    holiday_cols = []
    for c in day_block_cols:
        v = ws.cell(day_header_row, c).value
        if _is_text(v) and "FERIADO" in _norm(v):
            holiday_cols.append(c)
            # incluir columnas siguientes si el header está vacío pero la fila de fecha tiene algo
            for cc in range(c + 1, totals_start):
                hv = ws.cell(day_header_row, cc).value
                dv = ws.cell(day_header_row + 1, cc).value
                if (not _is_text(hv)) and (dv is not None and str(dv).strip() != ""):
                    holiday_cols.append(cc)
                else:
                    break
            break

    return day_header_row, dom_col, totals_start, holiday_cols


def extract_rows(ws, name_col, day_header_row, dom_col, totals_start, holiday_cols):
    """
    Extrae filas con empleados y valores desde una planilla generada.
    Heurística:
      - date_row = day_header_row + 1
      - data_row_start = primera fila con legajo + nombre debajo del header (busca hasta +25)
    """
    date_row = day_header_row + 1

    empresa_col = 1
    sector_col = 2
    legajo_col = 3

    # buscar inicio real de datos (en vez de asumir +4 fijo)
    data_row_start = None
    for r in range(day_header_row + 2, min(ws.max_row, day_header_row + 30) + 1):
        name = ws.cell(r, name_col).value
        leg = ws.cell(r, legajo_col).value
        name_ok = (name is not None and str(name).strip() != "")
        leg_ok = (leg is not None and str(leg).strip() != "")
        if name_ok and leg_ok:
            data_row_start = r
            break

    if data_row_start is None:
        # fallback: usar day_header_row+4 como antes
        data_row_start = day_header_row + 4

    day_labels = ["Dom", "Lun", "Mar", "Miérc", "Juev", "Vier", "Sáb"]
    day_cols = {day_labels[i]: dom_col + i for i in range(7)}

    holiday_dates = [ws.cell(date_row, c).value for c in holiday_cols]

    # detectar columnas de totales por texto del header
    totals_map = {}
    scan_row = day_header_row
    for c in range(totals_start, min(ws.max_column, totals_start + 60) + 1):
        v = ws.cell(scan_row, c).value
        txt = _norm(v) if _is_text(v) else ""
        if ("$/H" in txt) and ("L A V" in txt or "L A V" in txt.replace("(", " ").replace(")", " ")):
            totals_map["total_lv"] = c
        elif ("$/H" in txt) and ("SÁB" in txt or "SAB" in txt):
            totals_map["total_sab"] = c
        elif ("$/H" in txt) and ("DOM" in txt or "FER" in txt):
            totals_map["total_domfer"] = c
        elif "SUB TOTAL" in txt or "SUBTOTAL" in txt:
            totals_map["subtotal"] = c
        elif "REDONDE" in txt:
            totals_map["redondeo"] = c

    rows = []
    empty_streak = 0
    r = data_row_start

    while r <= ws.max_row and r <= data_row_start + 5000:
        name = ws.cell(r, name_col).value
        leg = ws.cell(r, legajo_col).value
        emp = ws.cell(r, empresa_col).value
        sec = ws.cell(r, sector_col).value

        name_ok = (name is not None and str(name).strip() != "")
        leg_ok = (leg is not None and str(leg).strip() != "")

        if not name_ok and not leg_ok:
            empty_streak += 1
            if empty_streak >= 12:
                break
            r += 1
            continue

        empty_streak = 0

        item = {
            "empresa": emp,
            "sector": sec,
            "legajo": leg,
            "nombre": name,
            "days": {k: ws.cell(r, c).value for k, c in day_cols.items()},
            "holidays": [ws.cell(r, c).value for c in holiday_cols],
            "totals": {
                "total_lv": ws.cell(r, totals_map["total_lv"]).value if totals_map.get("total_lv") else None,
                "total_sab": ws.cell(r, totals_map["total_sab"]).value if totals_map.get("total_sab") else None,
                "total_domfer": ws.cell(r, totals_map["total_domfer"]).value if totals_map.get("total_domfer") else None,
                "subtotal": ws.cell(r, totals_map["subtotal"]).value if totals_map.get("subtotal") else None,
                "redondeo": ws.cell(r, totals_map["redondeo"]).value if totals_map.get("redondeo") else None,
            }
        }
        rows.append(item)
        r += 1

    rows.sort(key=lambda x: _norm(str(x.get("nombre", ""))))

    day_dates = {lab: ws.cell(date_row, dom_col + i).value for i, lab in enumerate(day_labels)}

    return day_dates, holiday_dates, rows


def write_print_sheet(ws, day_dates, holiday_dates, rows):
    """
    Escribe una hoja lista para impresión.
    """
    thick = Side(style="thick")
    thin = Side(style="thin")
    b_thick = Border(left=thick, right=thick, top=thick, bottom=thick)
    b_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    font_h = Font(name="Calibri", size=11, bold=True)
    font_s = Font(name="Calibri", size=10, bold=True)
    font_b = Font(name="Calibri", size=10)

    yellow = PatternFill("solid", fgColor="FFF200")

    header_row = 3
    date_row = 4
    hs_row = 5
    price_row = 6
    data_start = 7

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 30

    fixed_headers = [("A", "EMPRESA"), ("B", "SECTOR"), ("C", "LEGAJO"), ("D", "APELLIDO Y\nNOMBRE")]
    for col, text in fixed_headers:
        ws.merge_cells(f"{col}{header_row}:{col}{price_row}")
        cell = ws[f"{col}{header_row}"]
        cell.value = text
        cell.font = font_h
        cell.alignment = center
        cell.border = b_thick

    day_labels = [("E", "Dom"), ("F", "Lun"), ("G", "Mar"), ("H", "Miérc"), ("I", "Juev"), ("J", "Vier"), ("K", "Sáb")]
    for col, lab in day_labels:
        ws.column_dimensions[col].width = 10

        ws[f"{col}{header_row}"].value = lab
        ws[f"{col}{header_row}"].font = font_h
        ws[f"{col}{header_row}"].alignment = center
        ws[f"{col}{header_row}"].border = b_thick

        ws[f"{col}{date_row}"].value = day_dates.get(lab)
        ws[f"{col}{date_row}"].font = font_s
        ws[f"{col}{date_row}"].alignment = center
        ws[f"{col}{date_row}"].border = b_thick

        ws[f"{col}{hs_row}"].value = "Hs"
        ws[f"{col}{hs_row}"].font = font_s
        ws[f"{col}{hs_row}"].alignment = center
        ws[f"{col}{hs_row}"].border = b_thick

        ws[f"{col}{price_row}"].value = ""
        ws[f"{col}{price_row}"].border = b_thick

    if not holiday_dates:
        holiday_dates = [""]

    hol_start_col_idx = 12  # L
    hol_cols = [get_column_letter(hol_start_col_idx + i) for i in range(len(holiday_dates))]

    ws.merge_cells(f"{hol_cols[0]}{header_row}:{hol_cols[-1]}{header_row}")
    ch = ws[f"{hol_cols[0]}{header_row}"]
    ch.value = "FERIADO"
    ch.font = font_h
    ch.alignment = center
    ch.border = b_thick

    for i, col in enumerate(hol_cols):
        ws.column_dimensions[col].width = 10

        ws[f"{col}{date_row}"].value = holiday_dates[i]
        ws[f"{col}{date_row}"].font = font_s
        ws[f"{col}{date_row}"].alignment = center
        ws[f"{col}{date_row}"].border = b_thick

        ws[f"{col}{hs_row}"].value = "Hs"
        ws[f"{col}{hs_row}"].font = font_s
        ws[f"{col}{hs_row}"].alignment = center
        ws[f"{col}{hs_row}"].border = b_thick

        ws[f"{col}{price_row}"].value = ""
        ws[f"{col}{price_row}"].border = b_thick

    base = hol_start_col_idx + len(hol_cols)
    cols_tot = {
        "total_lv": get_column_letter(base + 0),
        "total_sab": get_column_letter(base + 1),
        "total_domfer": get_column_letter(base + 2),
        "subtotal": get_column_letter(base + 3),
        "redondeo": get_column_letter(base + 4),
    }
    tot_headers = {
        "total_lv": "$/h (L a V)",
        "total_sab": "$/h (Sábado)",
        "total_domfer": "$/h (Dom y Fer)",
        "subtotal": "Sub Total",
        "redondeo": "Redondeo",
    }

    for key, col in cols_tot.items():
        ws.column_dimensions[col].width = 14

        if key in ("subtotal", "redondeo"):
            ws.merge_cells(f"{col}{header_row}:{col}{price_row}")
            cell = ws[f"{col}{header_row}"]
            cell.value = tot_headers[key]
            cell.font = font_h
            cell.alignment = center
            cell.border = b_thick
        else:
            ws.merge_cells(f"{col}{header_row}:{col}{hs_row}")
            cell = ws[f"{col}{header_row}"]
            cell.value = tot_headers[key]
            cell.font = font_h
            cell.alignment = center
            cell.border = b_thick

            ws[f"{col}{price_row}"].value = "precio"
            ws[f"{col}{price_row}"].fill = yellow
            ws[f"{col}{price_row}"].font = font_s
            ws[f"{col}{price_row}"].alignment = center
            ws[f"{col}{price_row}"].border = b_thick

    order_days = ["Dom", "Lun", "Mar", "Miérc", "Juev", "Vier", "Sáb"]

    for i, item in enumerate(rows):
        r = data_start + i

        ws[f"A{r}"].value = item.get("empresa")
        ws[f"B{r}"].value = item.get("sector")
        ws[f"C{r}"].value = item.get("legajo")
        ws[f"D{r}"].value = item.get("nombre")

        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{r}"].font = font_b
            ws[f"{col}{r}"].alignment = left if col == "D" else center
            ws[f"{col}{r}"].border = b_thin

        for j, dlab in enumerate(order_days):
            col = get_column_letter(5 + j)
            v = item.get("days", {}).get(dlab)
            ws[f"{col}{r}"].value = 0 if v in (None, "", "0") else v
            ws[f"{col}{r}"].font = font_b
            ws[f"{col}{r}"].alignment = center
            ws[f"{col}{r}"].border = b_thin

        for j, col in enumerate(hol_cols):
            v = item.get("holidays", [])[j] if j < len(item.get("holidays", [])) else 0
            ws[f"{col}{r}"].value = 0 if v in (None, "", "0") else v
            ws[f"{col}{r}"].font = font_b
            ws[f"{col}{r}"].alignment = center
            ws[f"{col}{r}"].border = b_thin

        ws[f"{cols_tot['total_lv']}{r}"].value = item.get("totals", {}).get("total_lv")
        ws[f"{cols_tot['total_sab']}{r}"].value = item.get("totals", {}).get("total_sab")
        ws[f"{cols_tot['total_domfer']}{r}"].value = item.get("totals", {}).get("total_domfer")
        ws[f"{cols_tot['subtotal']}{r}"].value = item.get("totals", {}).get("subtotal")
        ws[f"{cols_tot['redondeo']}{r}"].value = item.get("totals", {}).get("redondeo")

        money_cols = [cols_tot["total_lv"], cols_tot["total_sab"], cols_tot["total_domfer"], cols_tot["subtotal"], cols_tot["redondeo"]]
        for col in money_cols:
            ws[f"{col}{r}"].font = font_b
            ws[f"{col}{r}"].alignment = center
            ws[f"{col}{r}"].border = b_thin
            ws[f"{col}{r}"].number_format = '"$"#,##0'

    last_row = data_start + len(rows) - 1 if rows else data_start
    last_col = cols_tot["redondeo"]
    ws.print_area = f"A1:{last_col}{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.freeze_panes = f"A{data_start}"

    ws.row_dimensions[header_row].height = 20
    ws.row_dimensions[date_row].height = 18
    ws.row_dimensions[hs_row].height = 16
    ws.row_dimensions[price_row].height = 16


def build_rrhh_print_workbook(path_oeste, path_consultora, out_path):
    """
    Genera un archivo de impresión con 2 hojas:
    - OESTE
    - CONSULTORA
    """
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for label, path in [("OESTE", path_oeste), ("CONSULTORA", path_consultora)]:
        wb_in = load_workbook(path, data_only=False)  # conservar fórmulas si existen

        ws_src, header_row_name, name_col = find_table_sheet_and_header_row(wb_in)
        day_header_row, dom_col, totals_start, holiday_cols = find_day_and_totals_columns(ws_src, header_row_name)
        day_dates, holiday_dates, rows = extract_rows(ws_src, name_col, day_header_row, dom_col, totals_start, holiday_cols)

        ws_out = wb_out.create_sheet(label)
        write_print_sheet(ws_out, day_dates, holiday_dates, rows)

    wb_out.save(out_path)
